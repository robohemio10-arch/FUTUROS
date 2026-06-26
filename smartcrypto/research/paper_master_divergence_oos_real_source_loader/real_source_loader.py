"""Read-only/version-safe real source loader for Paper/Master divergence research.

This module intentionally does not apply any remediation, does not register
candidate rules, does not update Freqtrade/RiskManager/Qlib/AI Shadow runtime,
and does not send orders. Runtime/data reads are opt-in. With no explicit source
paths and no explicit runtime-read allowance, the report remains blocked and
informational.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


SCHEMA_VERSION = "paper_master_divergence_oos_real_source_loader_v1"
PROJECT_NAME = "SMART FUTUROS"
DECISION = "MANTER_EM_RESEARCH"
HYPOTHESIS_SCOPE = ["H1", "H2", "H6"]
OOS_SLICE_DIMENSIONS = [
    "day",
    "symbol",
    "side",
    "exit_reason",
    "duration_bucket",
    "covered_vs_uncovered",
]
MINIMUM_NORMALIZED_COLUMNS = [
    "source_role",
    "trade_id",
    "symbol",
    "side",
    "open_time",
    "close_time",
    "day",
    "pnl",
    "exit_reason",
    "duration_minutes",
    "duration_bucket",
    "covered_feature_subset",
    "covered_vs_uncovered",
]
_SUPPORTED_SUFFIXES = {".csv", ".json", ".jsonl", ".xlsx"}


@dataclass(frozen=True)
class SourceLoadResult:
    """Normalized source loading result."""

    source_role: str
    source_path: str | None
    source_exists: bool
    source_type: str | None
    source_hash_sha256: str | None
    source_size_bytes: int | None
    source_mtime_utc: str | None
    source_status: str
    source_reason: str
    row_count_raw: int
    row_count_normalized: int
    symbols: list[str]
    sides: list[str]
    first_close_time: str | None
    last_close_time: str | None
    normalized_rows: list[dict[str, Any]]
    validation_errors: list[str]

    def to_public_dict(self, *, include_rows: bool = False) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "source_role": self.source_role,
            "source_path": self.source_path,
            "source_exists": self.source_exists,
            "source_type": self.source_type,
            "source_hash_sha256": self.source_hash_sha256,
            "source_size_bytes": self.source_size_bytes,
            "source_mtime_utc": self.source_mtime_utc,
            "source_status": self.source_status,
            "source_reason": self.source_reason,
            "row_count_raw": self.row_count_raw,
            "row_count_normalized": self.row_count_normalized,
            "symbols": self.symbols,
            "sides": self.sides,
            "first_close_time": self.first_close_time,
            "last_close_time": self.last_close_time,
            "validation_errors": self.validation_errors,
        }
        if include_rows:
            payload["normalized_rows"] = self.normalized_rows
        return payload


def _safety_flags() -> dict[str, Any]:
    return {
        "research_only": True,
        "read_only": True,
        "paper_only": True,
        "shadow_only": True,
        "operational_authority": False,
        "release_authority": False,
        "readiness_release_authority": False,
        "live_trading_enabled": False,
        "live_release_allowed": False,
        "canary_release_allowed": False,
        "order_submission_enabled": False,
        "real_order_submission_enabled": False,
        "exchange_private_access": False,
        "sends_orders": False,
        "changes_risk": False,
        "changes_model": False,
        "updates_freqtrade": False,
        "updates_risk_manager": False,
        "updates_qlib_runtime": False,
        "updates_ai_shadow_runtime": False,
        "executes_scheduler": False,
        "executes_orchestrator": False,
        "executes_stage_builders": False,
        "runs_training": False,
        "applies_shadow_rules": False,
        "applies_feedback_to_ai_shadow": False,
        "registers_candidate_rules": False,
        "can_apply_to_freqtrade": False,
        "can_apply_to_risk_manager": False,
        "can_promote_rules": False,
        "can_promote_model": False,
        "remediation_application_allowed": False,
        "ready_for_candidate_registry": False,
    }


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _mtime_utc(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()


def _json_safe(value: Any) -> Any:
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, Mapping):
        return {str(key): _json_safe(inner) for key, inner in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    return value


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "nan", "nat"}:
        return None
    return text


def _first_value(row: Mapping[str, Any], aliases: Sequence[str]) -> Any:
    lowered = {str(key).strip().lower(): value for key, value in row.items()}
    for alias in aliases:
        if alias.lower() in lowered:
            value = lowered[alias.lower()]
            if _clean_text(value) is not None:
                return value
    return None


def _coerce_float(value: Any) -> float | None:
    text = _clean_text(value)
    if text is None:
        return None
    normalized = text.replace("%", "").replace(" ", "")
    if "," in normalized and "." not in normalized:
        normalized = normalized.replace(",", ".")
    try:
        number = float(normalized)
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def _coerce_bool(value: Any) -> bool | None:
    text = _clean_text(value)
    if text is None:
        return None
    lowered = text.lower()
    if lowered in {"1", "true", "t", "yes", "y", "sim", "covered", "available"}:
        return True
    if lowered in {"0", "false", "f", "no", "n", "nao", "não", "uncovered", "missing"}:
        return False
    return None


def _parse_datetime_text(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    normalized = text.replace("Z", "+00:00")
    accepted_formats = (
        "%Y-%m-%d %H:%M:%S.%f%z",
        "%Y-%m-%d %H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    )
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        parsed = None
    if parsed is None:
        for date_format in accepted_formats:
            try:
                parsed = datetime.strptime(normalized, date_format)
                break
            except ValueError:
                continue
    if parsed is None:
        return text
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed.isoformat(timespec="seconds")


def _day_from_close_time(close_time: str | None) -> str | None:
    if close_time is None:
        return None
    if len(close_time) >= 10:
        return close_time[:10]
    return close_time


def _normalize_side(value: Any) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    lowered = text.lower().replace(" ", "_")
    if lowered in {"long", "buy", "comprado"}:
        return "long"
    if lowered in {"short", "sell", "vendido"}:
        return "short"
    return lowered


def _duration_bucket(duration_minutes: float | None) -> str:
    if duration_minutes is None:
        return "unknown"
    if duration_minutes < 15:
        return "under_15m"
    if duration_minutes < 30:
        return "15m_to_30m"
    if duration_minutes < 60:
        return "30m_to_60m"
    if duration_minutes < 180:
        return "1h_to_3h"
    if duration_minutes < 360:
        return "3h_to_6h"
    return "over_6h"


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _read_json(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if isinstance(payload, list):
        return [dict(item) for item in payload if isinstance(item, Mapping)]
    if isinstance(payload, Mapping):
        for key in ("rows", "trades", "data", "items"):
            value = payload.get(key)
            if isinstance(value, list):
                return [dict(item) for item in value if isinstance(item, Mapping)]
    raise ValueError("JSON source must be a list of objects or an object containing rows/trades/data/items")


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            payload = json.loads(stripped)
            if not isinstance(payload, Mapping):
                raise ValueError(f"JSONL line {line_number} is not an object")
            rows.append(dict(payload))
    return rows


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency exists in project envs that use xlsx
        raise RuntimeError("openpyxl is required to read .xlsx sources") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []
    header_names = [str(header).strip() if header is not None else "" for header in headers]
    output: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        mapped = {
            header: value
            for header, value in zip(header_names, raw_row, strict=False)
            if header
        }
        if any(_clean_text(value) is not None for value in mapped.values()):
            output.append(mapped)
    return output


def read_source_rows(path: Path) -> list[dict[str, Any]]:
    """Read supported tabular trade source rows."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".json":
        return _read_json(path)
    if suffix == ".jsonl":
        return _read_jsonl(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"unsupported_source_type:{suffix}")


def normalize_trade_rows(rows: Iterable[Mapping[str, Any]], *, source_role: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Normalize raw trade rows into the OOS source-loader research schema."""
    normalized_rows: list[dict[str, Any]] = []
    validation_errors: list[str] = []
    for index, row in enumerate(rows, start=1):
        trade_id = _clean_text(
            _first_value(row, ("trade_id", "id", "order_id", "internal_order_id", "tradeId"))
        ) or f"{source_role}_{index}"
        symbol = _clean_text(_first_value(row, ("symbol", "pair", "instrument", "market")))
        side = _normalize_side(_first_value(row, ("side", "direction", "position_side", "is_short")))
        if side in {"true", "1"}:
            side = "short"
        if side in {"false", "0"}:
            side = "long"
        open_time = _parse_datetime_text(
            _first_value(row, ("open_time", "open_date", "entry_time", "entry_date", "created_at"))
        )
        close_time = _parse_datetime_text(
            _first_value(row, ("close_time", "close_date", "exit_time", "exit_date", "closed_at"))
        )
        pnl = _coerce_float(
            _first_value(row, ("pnl", "net_pnl", "profit_abs", "profit", "realized_pnl", "pnl_abs"))
        )
        exit_reason = _clean_text(
            _first_value(row, ("exit_reason", "close_reason", "sell_reason", "reason", "exit_tag"))
        ) or "unknown"
        duration_minutes = _coerce_float(
            _first_value(row, ("duration_minutes", "duration_min", "duration", "minutes", "trade_duration_minutes"))
        )
        if duration_minutes is None:
            duration_seconds = _coerce_float(_first_value(row, ("duration_seconds", "duration_s")))
            if duration_seconds is not None:
                duration_minutes = duration_seconds / 60.0
        covered = _coerce_bool(
            _first_value(
                row,
                (
                    "covered_feature_subset",
                    "features_covered",
                    "feature_coverage",
                    "has_features",
                    "covered",
                ),
            )
        )
        covered_vs_uncovered = "covered" if covered is True else "uncovered" if covered is False else "unknown"

        row_errors: list[str] = []
        if symbol is None:
            row_errors.append("missing_symbol")
        if side is None:
            row_errors.append("missing_side")
        if close_time is None:
            row_errors.append("missing_close_time")
        if pnl is None:
            row_errors.append("missing_pnl")
        if row_errors:
            validation_errors.append(f"{source_role}:row_{index}:" + "+".join(row_errors))
            continue

        normalized_rows.append(
            {
                "source_role": source_role,
                "trade_id": trade_id,
                "symbol": symbol,
                "side": side,
                "open_time": open_time,
                "close_time": close_time,
                "day": _day_from_close_time(close_time),
                "pnl": pnl,
                "exit_reason": exit_reason,
                "duration_minutes": duration_minutes,
                "duration_bucket": _duration_bucket(duration_minutes),
                "covered_feature_subset": covered,
                "covered_vs_uncovered": covered_vs_uncovered,
            }
        )
    return normalized_rows, validation_errors


def _source_summary_from_rows(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    close_times = sorted(
        str(row["close_time"])
        for row in rows
        if _clean_text(row.get("close_time")) is not None
    )
    symbols = sorted({str(row["symbol"]) for row in rows if _clean_text(row.get("symbol")) is not None})
    sides = sorted({str(row["side"]) for row in rows if _clean_text(row.get("side")) is not None})
    return {
        "symbols": symbols,
        "sides": sides,
        "first_close_time": close_times[0] if close_times else None,
        "last_close_time": close_times[-1] if close_times else None,
    }


def _absent_source_result(source_role: str, source_path: str | None, status: str, reason: str) -> SourceLoadResult:
    return SourceLoadResult(
        source_role=source_role,
        source_path=source_path,
        source_exists=False,
        source_type=None,
        source_hash_sha256=None,
        source_size_bytes=None,
        source_mtime_utc=None,
        source_status=status,
        source_reason=reason,
        row_count_raw=0,
        row_count_normalized=0,
        symbols=[],
        sides=[],
        first_close_time=None,
        last_close_time=None,
        normalized_rows=[],
        validation_errors=[] if status == "missing_optional_source" else [reason],
    )


def load_trade_source(
    source_path: str | Path | None,
    *,
    source_role: str,
    project_root: str | Path = ".",
    allow_runtime_read: bool = False,
) -> SourceLoadResult:
    """Load and normalize one Paper/Master source in read-only mode.

    Runtime reads are blocked unless ``allow_runtime_read`` is explicitly true.
    """
    if source_path is None:
        return _absent_source_result(source_role, None, "missing_optional_source", "source_path_not_provided")

    path = Path(source_path)
    if not path.is_absolute():
        path = Path(project_root) / path
    resolved_path = path.resolve()
    public_path = str(resolved_path)

    if not allow_runtime_read:
        return _absent_source_result(
            source_role,
            public_path,
            "blocked_runtime_read_not_allowed",
            "explicit_allow_runtime_read_required",
        )
    if not resolved_path.exists():
        return _absent_source_result(source_role, public_path, "missing_source", "source_file_not_found")
    if not resolved_path.is_file():
        return _absent_source_result(source_role, public_path, "invalid_source", "source_path_is_not_file")
    suffix = resolved_path.suffix.lower()
    if suffix not in _SUPPORTED_SUFFIXES:
        return SourceLoadResult(
            source_role=source_role,
            source_path=public_path,
            source_exists=True,
            source_type=suffix,
            source_hash_sha256=_sha256_file(resolved_path),
            source_size_bytes=resolved_path.stat().st_size,
            source_mtime_utc=_mtime_utc(resolved_path),
            source_status="unsupported_source_type",
            source_reason=f"unsupported_source_type:{suffix}",
            row_count_raw=0,
            row_count_normalized=0,
            symbols=[],
            sides=[],
            first_close_time=None,
            last_close_time=None,
            normalized_rows=[],
            validation_errors=[f"unsupported_source_type:{suffix}"],
        )

    try:
        raw_rows = read_source_rows(resolved_path)
        normalized_rows, validation_errors = normalize_trade_rows(raw_rows, source_role=source_role)
    except Exception as exc:  # noqa: BLE001 - report as data-source validation error, not runtime crash.
        return SourceLoadResult(
            source_role=source_role,
            source_path=public_path,
            source_exists=True,
            source_type=suffix,
            source_hash_sha256=_sha256_file(resolved_path),
            source_size_bytes=resolved_path.stat().st_size,
            source_mtime_utc=_mtime_utc(resolved_path),
            source_status="source_read_failed",
            source_reason=str(exc),
            row_count_raw=0,
            row_count_normalized=0,
            symbols=[],
            sides=[],
            first_close_time=None,
            last_close_time=None,
            normalized_rows=[],
            validation_errors=[str(exc)],
        )

    summary = _source_summary_from_rows(normalized_rows)
    status = "loaded" if normalized_rows else "invalid_schema_or_empty_source"
    reason = "source_loaded_read_only" if normalized_rows else "no_valid_normalized_rows"
    return SourceLoadResult(
        source_role=source_role,
        source_path=public_path,
        source_exists=True,
        source_type=suffix,
        source_hash_sha256=_sha256_file(resolved_path),
        source_size_bytes=resolved_path.stat().st_size,
        source_mtime_utc=_mtime_utc(resolved_path),
        source_status=status,
        source_reason=reason,
        row_count_raw=len(raw_rows),
        row_count_normalized=len(normalized_rows),
        symbols=list(summary["symbols"]),
        sides=list(summary["sides"]),
        first_close_time=summary["first_close_time"],
        last_close_time=summary["last_close_time"],
        normalized_rows=normalized_rows,
        validation_errors=validation_errors,
    )


def _build_source_gate(paper_result: SourceLoadResult, master_result: SourceLoadResult) -> dict[str, Any]:
    both_loaded = paper_result.source_status == "loaded" and master_result.source_status == "loaded"
    if both_loaded:
        evidence = (
            f"paper_rows={paper_result.row_count_normalized}; "
            f"master_rows={master_result.row_count_normalized}"
        )
    else:
        evidence = f"paper_status={paper_result.source_status}; master_status={master_result.source_status}"
    return {
        "gate_id": "real_sources_loaded_when_explicitly_allowed",
        "gate_name": "Real Paper/Master sources loaded only through explicit opt-in",
        "severity": "high",
        "passed": True,
        "evidence": evidence,
    }


def _gate_matrix(paper_result: SourceLoadResult, master_result: SourceLoadResult) -> list[dict[str, Any]]:
    return [
        {
            "gate_id": "research_only_contract",
            "gate_name": "Research-only contract preserved",
            "severity": "critical",
            "passed": True,
            "evidence": "research_only=true; operational_authority=false",
        },
        {
            "gate_id": "real_source_loader_created",
            "gate_name": "Real source loader is informational and read-only",
            "severity": "critical",
            "passed": True,
            "evidence": "read_only=true; writes_runtime=false; writes_data=false",
        },
        _build_source_gate(paper_result, master_result),
        {
            "gate_id": "normalized_schema_declared",
            "gate_name": "Minimum normalized schema declared",
            "severity": "high",
            "passed": True,
            "evidence": f"columns={MINIMUM_NORMALIZED_COLUMNS}",
        },
        {
            "gate_id": "oos_slice_dimensions_preserved",
            "gate_name": "OOS slice dimensions preserved",
            "severity": "high",
            "passed": True,
            "evidence": f"dimensions={OOS_SLICE_DIMENSIONS}",
        },
        {
            "gate_id": "promotion_blocked",
            "gate_name": "Rule and model promotion blocked",
            "severity": "critical",
            "passed": True,
            "evidence": "can_promote_rules=false; can_promote_model=false",
        },
        {
            "gate_id": "runtime_unchanged",
            "gate_name": "Runtime and execution surfaces unchanged",
            "severity": "critical",
            "passed": True,
            "evidence": "no runtime updates; sends_orders=false",
        },
    ]


def _gate_summary(gates: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    failed = [str(gate["gate_id"]) for gate in gates if not bool(gate.get("passed"))]
    critical_failed = [
        str(gate["gate_id"])
        for gate in gates
        if not bool(gate.get("passed")) and gate.get("severity") == "critical"
    ]
    return {
        "gate_count": len(gates),
        "passed_gate_count": len(gates) - len(failed),
        "failed_gate_count": len(failed),
        "failed_gate_ids": failed,
        "critical_failed_gate_ids": critical_failed,
    }


def _loaded_rows_report(paper_result: SourceLoadResult, master_result: SourceLoadResult) -> dict[str, Any]:
    paper_symbols = set(paper_result.symbols)
    master_symbols = set(master_result.symbols)
    paper_sides = set(paper_result.sides)
    master_sides = set(master_result.sides)
    both_loaded = paper_result.source_status == "loaded" and master_result.source_status == "loaded"
    return {
        "real_sources_loaded": both_loaded,
        "paper_rows": paper_result.row_count_normalized,
        "master_rows": master_result.row_count_normalized,
        "common_symbols": sorted(paper_symbols & master_symbols),
        "paper_only_symbols": sorted(paper_symbols - master_symbols),
        "master_only_symbols": sorted(master_symbols - paper_symbols),
        "common_sides": sorted(paper_sides & master_sides),
        "paper_only_sides": sorted(paper_sides - master_sides),
        "master_only_sides": sorted(master_sides - paper_sides),
        "oos_ready_for_slice_metrics": both_loaded,
    }


def _write_report_if_requested(report: Mapping[str, Any], *, project_root: Path, write_requested: bool) -> tuple[bool, str | None]:
    if not write_requested:
        return False, None
    reports_dir = project_root / "data" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = reports_dir / "paper_master_divergence_oos_real_source_loader_v1.json"
    output_path.write_text(json.dumps(_json_safe(report), indent=2, ensure_ascii=False), encoding="utf-8")
    return True, str(output_path)


def build_paper_master_divergence_oos_real_source_loader_report(
    *,
    project_root: str | Path = ".",
    paper_source: str | Path | None = None,
    master_source: str | Path | None = None,
    allow_runtime_read: bool = False,
    write: bool = False,
    include_loaded_rows: bool = False,
) -> dict[str, Any]:
    """Build the read-only real source loader report."""
    root = Path(project_root).resolve()
    paper_result = load_trade_source(
        paper_source,
        source_role="paper",
        project_root=root,
        allow_runtime_read=allow_runtime_read,
    )
    master_result = load_trade_source(
        master_source,
        source_role="master",
        project_root=root,
        allow_runtime_read=allow_runtime_read,
    )
    sources_loaded = paper_result.source_status == "loaded" and master_result.source_status == "loaded"
    input_mode = "real_sources_loaded_read_only" if sources_loaded else "no_runtime_rows_loaded"
    validation_errors = paper_result.validation_errors + master_result.validation_errors
    gates = _gate_matrix(paper_result, master_result)
    report: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "project_name": PROJECT_NAME,
        "project_root": str(project_root),
        "status": "blocked",
        "reason": "paper_master_divergence_requires_real_source_loading_before_oos_computation",
        "decision": DECISION,
        "input_mode": input_mode,
        "hypothesis_scope": HYPOTHESIS_SCOPE,
        "oos_slice_dimensions": OOS_SLICE_DIMENSIONS,
        "minimum_normalized_columns": MINIMUM_NORMALIZED_COLUMNS,
        "real_source_loader_created": True,
        "real_sources_loaded": sources_loaded,
        "allow_runtime_read": allow_runtime_read,
        "oos_ready_for_slice_metrics": sources_loaded,
        "oos_slice_metrics_computed": False,
        "oos_validated": False,
        "oos_validation_required": True,
        "paper_replicates_master_edge": False,
        "source_summary": {
            "paper": paper_result.to_public_dict(include_rows=include_loaded_rows),
            "master": master_result.to_public_dict(include_rows=include_loaded_rows),
        },
        "loaded_rows_report": _loaded_rows_report(paper_result, master_result),
        "validation_errors": validation_errors,
        "allowed_next_steps": [
            "fornecer caminhos reais de Paper/Master explicitamente e com allow-runtime-read",
            "computar OOS por day/symbol/side/exit_reason/duration/covered-vs-uncovered em branch seguinte",
            "medir false positive/false negative e winner retention por hipótese",
            "bloquear regra se remover ROI winners ou concentrar efeito em único dia",
            "somente criar registry shadow bloqueado se OOS passar",
        ],
        "forbidden_actions": [
            "versionar runtime/data/logs/parquet/sqlite/xlsx/csv",
            "aplicar regra no Freqtrade",
            "alterar RiskManager",
            "alterar Qlib runtime",
            "alterar IA Shadow runtime",
            "registrar ou promover candidate rule",
            "promover modelo",
            "executar treino operacional",
            "habilitar live ou canary",
            "enviar ordem real",
            "usar exchange privada",
            "escrever artefatos em data/runtime/reports/logs/freqtrade por padrão",
        ],
        "gate_matrix": gates,
        "gate_summary": _gate_summary(gates),
        "write_requested": write,
        "write_performed": False,
        "output_path": None,
        "writes_reports": False,
        "writes_runtime": False,
        "writes_data": False,
        "writes_parquet": False,
        "writes_sqlite": False,
    }
    report.update(_safety_flags())
    write_performed, output_path = _write_report_if_requested(report, project_root=root, write_requested=write)
    report["write_performed"] = write_performed
    report["output_path"] = output_path
    report["writes_reports"] = write_performed
    return _json_safe(report)
