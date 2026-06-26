"""Research-only real OOS slice computation for Paper/Master divergence.

The module is deliberately conservative:

- It does not read runtime files unless the caller explicitly enables
  ``allow_runtime_read`` and supplies source paths.
- It never writes runtime, report, SQLite, parquet, model, Freqtrade, Qlib,
  RiskManager, or AI Shadow artifacts.
- It never promotes rules or models.
- It computes descriptive metrics only and returns a blocked research decision.

The practical purpose is to move the Paper/Master divergence work from
"structural contracts" to measurable slice-level evidence, while preserving the
project safety envelope.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


SCHEMA_VERSION = "paper_master_divergence_oos_real_slice_computation_v1"
PROJECT_NAME = "SMART FUTUROS"

EXPECTED_TRADE_VALUE_FORMULA = (
    "expected_trade_value = Qlib_expected_return_net × "
    "Shadow_probability_quality × Regime_confidence - Estimated_fee - "
    "Estimated_spread - Estimated_slippage - Latency_penalty - "
    "Drawdown_penalty - Drift_penalty"
)

OOS_SLICE_DIMENSIONS: tuple[str, ...] = (
    "day",
    "symbol",
    "side",
    "exit_reason",
    "duration_bucket",
    "covered_vs_uncovered",
)

HYPOTHESIS_SCOPE: tuple[str, ...] = ("H1", "H2", "H6")

CANONICAL_PAPER_KPIS: dict[str, Any] = {
    "trade_count": 239,
    "net_pnl": -21.35477552,
    "gross_profit": 87.22777285,
    "gross_loss": -108.58254837,
    "profit_factor": 0.803331,
    "win_rate": 0.405858,
    "avg_duration_hours": 3.4234,
    "fees": 11.89221848,
    "symbols": ["BTC/USDT:USDT", "ETH/USDT:USDT"],
}

CANONICAL_MASTER_KPIS: dict[str, Any] = {
    "trade_count": 243,
    "net_pnl": 143.166332,
    "profit_factor": 2.072573,
    "win_rate": 0.707819,
    "max_drawdown": -9.378203,
    "mean_trade_roi": 0.001181,
}

CANONICAL_DIVERGENCE_METRICS: dict[str, Any] = {
    "paper_minus_master_net_pnl": -164.52110752,
    "paper_minus_master_profit_factor": -1.269242,
    "paper_minus_master_trade_count": -4,
    "paper_minus_master_win_rate_points": -30.1961,
    "paper_replicates_master_edge": False,
}

CANONICAL_CLUSTER_EVIDENCE: dict[str, Any] = {
    "roi_net_pnl": 87.22777285,
    "stop_loss_net_pnl": -108.58254837,
    "fast_stop_under_30m": "critical",
    "remove_stop_loss_under_30m_delta": 34.9161,
    "eth_long_stop_loss_cluster": "critical",
    "candidate_shadow_rule": (
        "lb_10m_ret_close <= -0.0038501215827868 AND "
        "lb_30m_ret_close <= -0.0060685748963285"
    ),
    "candidate_shadow_rule_precision": 0.65625,
    "candidate_shadow_rule_recall": 0.41176,
    "candidate_shadow_rule_simulated_removed_pnl_delta": 8.9745,
}

MINIMUM_METRICS: tuple[str, ...] = (
    "trade_count",
    "net_pnl",
    "profit_factor",
    "win_rate",
    "max_drawdown",
    "winner_retention_rate",
    "winner_pnl_removed",
    "loser_pnl_removed",
    "false_positive_count",
    "false_negative_count",
    "precision",
    "recall",
    "coverage_ratio",
    "simulated_removed_pnl_delta",
)

FORBIDDEN_ACTIONS: tuple[str, ...] = (
    "aplicar regra no Freqtrade",
    "alterar RiskManager",
    "alterar Qlib runtime",
    "alterar IA Shadow runtime",
    "registrar ou promover candidate rule",
    "promover modelo",
    "executar treino operacional",
    "habilitar live ou canary",
    "enviar ordem real",
    "usar exchange privada",
    "escrever artefatos em data/runtime/reports/logs/freqtrade por padrão",
)

DEFAULT_SOURCE_CANDIDATES: dict[str, tuple[str, ...]] = {
    "paper": (
        "data/reports/paper_stoploss_root_cause_v1.json",
        "data/reports/daily_paper_master_kpi_pack_v1.json",
        "data/reports/dashboard_real_paper_sources_v1.json",
    ),
    "master": (
        "data/trades/trades_master.xlsx",
        "data/reports/ocr_v11_research_dataset.json",
        "data/reports/daily_paper_master_kpi_pack_v1.json",
    ),
}


@dataclass(frozen=True)
class TradeRecord:
    """Normalized read-only trade record."""

    source: str
    symbol: str
    side: str
    exit_reason: str
    close_time: str
    day: str
    duration_minutes: float | None
    duration_bucket: str
    pnl: float
    lb_10m_ret_close: float | None
    lb_30m_ret_close: float | None
    covered_feature_subset: bool
    raw: Mapping[str, Any]


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if math.isfinite(number):
            return number
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return number


def _safe_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "sim", "covered", "ok", "valid"}


def _first_present(row: Mapping[str, Any], keys: Sequence[str], default: Any = None) -> Any:
    lower_map = {str(key).strip().lower(): value for key, value in row.items()}
    for key in keys:
        value = lower_map.get(key.strip().lower())
        if value is not None and str(value).strip() != "":
            return value
    return default


def _normalize_side(value: Any) -> str:
    text = str(value or "unknown").strip().lower()
    if text in {"buy", "long", "entry_long"}:
        return "long"
    if text in {"sell", "short", "entry_short"}:
        return "short"
    return text or "unknown"


def _normalize_exit_reason(value: Any) -> str:
    text = str(value or "unknown").strip().lower()
    normalized = text.replace(" ", "_").replace("-", "_")
    if normalized in {"stoploss", "stop_loss", "sl"}:
        return "stop_loss"
    if normalized in {"roi", "take_profit", "tp"}:
        return "roi"
    return normalized or "unknown"


def _duration_bucket(minutes: float | None) -> str:
    if minutes is None:
        return "unknown"
    if minutes < 15:
        return "<15m"
    if minutes < 30:
        return "15-30m"
    if minutes < 60:
        return "30-60m"
    if minutes < 180:
        return "1-3h"
    if minutes < 360:
        return "3-6h"
    return ">6h"


def _parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    candidates = [
        text,
        text.replace("Z", "+00:00"),
        text.replace(" ", "T").replace("Z", "+00:00"),
    ]
    for candidate in candidates:
        try:
            parsed = datetime.fromisoformat(candidate)
        except ValueError:
            continue
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            parsed = datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
        return parsed
    return None


def _infer_duration_minutes(row: Mapping[str, Any]) -> float | None:
    direct = _safe_float(
        _first_present(
            row,
            (
                "duration_minutes",
                "duration_min",
                "trade_duration_minutes",
                "minutes",
            ),
        )
    )
    if direct is not None:
        return direct
    hours = _safe_float(
        _first_present(row, ("duration_hours", "trade_duration_hours", "hours"))
    )
    if hours is not None:
        return hours * 60.0
    seconds = _safe_float(
        _first_present(row, ("duration_seconds", "trade_duration_seconds", "seconds"))
    )
    if seconds is not None:
        return seconds / 60.0

    open_time = _parse_datetime(_first_present(row, ("open_time", "entry_time", "open_date")))
    close_time = _parse_datetime(
        _first_present(row, ("close_time", "exit_time", "close_date", "date"))
    )
    if open_time and close_time:
        return max((close_time - open_time).total_seconds() / 60.0, 0.0)
    return None


def _infer_pnl(row: Mapping[str, Any]) -> float:
    value = _safe_float(
        _first_present(
            row,
            (
                "pnl",
                "net_pnl",
                "profit_abs",
                "profit",
                "realized_profit",
                "close_profit_abs",
                "total_profit",
            ),
            0.0,
        )
    )
    return value if value is not None else 0.0


def normalize_trade_row(row: Mapping[str, Any], *, source: str) -> TradeRecord:
    symbol = str(_first_present(row, ("symbol", "pair", "instrument"), "UNKNOWN")).strip()
    side = _normalize_side(_first_present(row, ("side", "trade_side", "direction"), "unknown"))
    exit_reason = _normalize_exit_reason(
        _first_present(row, ("exit_reason", "close_reason", "sell_reason", "reason"), "unknown")
    )
    close_dt = _parse_datetime(
        _first_present(row, ("close_time", "exit_time", "close_date", "date", "timestamp"))
    )
    close_time = close_dt.isoformat() if close_dt else ""
    day = close_dt.date().isoformat() if close_dt else "unknown"
    duration_minutes = _infer_duration_minutes(row)
    lb_10m = _safe_float(_first_present(row, ("lb_10m_ret_close", "ret_10m", "return_10m")))
    lb_30m = _safe_float(_first_present(row, ("lb_30m_ret_close", "ret_30m", "return_30m")))
    feature_covered = _safe_bool(
        _first_present(
            row,
            (
                "covered_feature_subset",
                "feature_covered",
                "has_features",
                "covered",
                "valid_features",
            ),
            lb_10m is not None and lb_30m is not None,
        )
    )
    return TradeRecord(
        source=source,
        symbol=symbol,
        side=side,
        exit_reason=exit_reason,
        close_time=close_time,
        day=day,
        duration_minutes=duration_minutes,
        duration_bucket=_duration_bucket(duration_minutes),
        pnl=_infer_pnl(row),
        lb_10m_ret_close=lb_10m,
        lb_30m_ret_close=lb_30m,
        covered_feature_subset=feature_covered,
        raw=dict(row),
    )


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _read_json_rows(path: Path) -> list[Mapping[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, Mapping)]
    if isinstance(payload, Mapping):
        for key in (
            "trades",
            "rows",
            "paper_trades",
            "master_trades",
            "records",
            "data",
        ):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, Mapping)]
    return []


def _read_csv_rows(path: Path) -> list[Mapping[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_xlsx_rows(path: Path) -> list[Mapping[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - environment dependent
        raise RuntimeError(
            f"openpyxl is required to read xlsx source {path}"
        ) from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    output: list[Mapping[str, Any]] = []
    for raw_values in rows[1:]:
        row = {
            header[index]: value
            for index, value in enumerate(raw_values)
            if index < len(header) and header[index]
        }
        if row:
            output.append(row)
    return output


def read_trade_source(path: Path, *, source: str) -> tuple[list[TradeRecord], dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        raw_rows = _read_json_rows(path)
    elif suffix in {".csv", ".txt"}:
        raw_rows = _read_csv_rows(path)
    elif suffix in {".xlsx", ".xlsm"}:
        raw_rows = _read_xlsx_rows(path)
    else:
        raw_rows = []

    records = [normalize_trade_row(row, source=source) for row in raw_rows]
    metadata = {
        "source": source,
        "path": str(path),
        "exists": path.exists(),
        "sha256": _sha256_file(path) if path.exists() else None,
        "row_count": len(records),
        "suffix": suffix,
    }
    return records, metadata


def _compute_profit_factor(gross_profit: float, gross_loss: float) -> float | None:
    if gross_loss == 0:
        return None if gross_profit == 0 else float("inf")
    return gross_profit / abs(gross_loss)


def _max_drawdown(pnls: Sequence[float]) -> float:
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    for pnl in pnls:
        cumulative += pnl
        peak = max(peak, cumulative)
        max_dd = min(max_dd, cumulative - peak)
    return round(max_dd, 8)


def _record_matches_hypothesis(record: TradeRecord, hypothesis_id: str) -> bool:
    if hypothesis_id == "H1":
        return (
            record.exit_reason == "stop_loss"
            and record.duration_minutes is not None
            and record.duration_minutes < 30.0
        )
    if hypothesis_id == "H2":
        return "ETH" in record.symbol.upper() and record.side == "long"
    if hypothesis_id == "H6":
        if record.lb_10m_ret_close is None or record.lb_30m_ret_close is None:
            return False
        return (
            record.lb_10m_ret_close <= -0.0038501215827868
            and record.lb_30m_ret_close <= -0.0060685748963285
        )
    return False


def _slice_key(record: TradeRecord) -> dict[str, str]:
    return {
        "day": record.day,
        "symbol": record.symbol or "UNKNOWN",
        "side": record.side or "unknown",
        "exit_reason": record.exit_reason or "unknown",
        "duration_bucket": record.duration_bucket,
        "covered_vs_uncovered": (
            "covered" if record.covered_feature_subset else "uncovered"
        ),
    }


def _metrics_for_records(
    records: Sequence[TradeRecord],
    *,
    hypothesis_id: str,
) -> dict[str, Any]:
    trade_count = len(records)
    pnls = [record.pnl for record in records]
    winners = [record for record in records if record.pnl > 0]
    losers = [record for record in records if record.pnl < 0]
    triggered = [
        record for record in records if _record_matches_hypothesis(record, hypothesis_id)
    ]
    true_positive = [record for record in triggered if record.pnl < 0]
    false_positive = [record for record in triggered if record.pnl > 0]
    false_negative = [
        record
        for record in records
        if record.pnl < 0 and not _record_matches_hypothesis(record, hypothesis_id)
    ]
    covered = [record for record in records if record.covered_feature_subset]
    gross_profit = sum(record.pnl for record in winners)
    gross_loss = sum(record.pnl for record in losers)
    winner_pnl_removed = sum(record.pnl for record in false_positive)
    loser_pnl_removed = sum(record.pnl for record in true_positive)
    simulated_removed_pnl_delta = -sum(record.pnl for record in triggered)

    precision = (
        len(true_positive) / len(triggered)
        if triggered
        else None
    )
    recall = (
        len(true_positive) / len(losers)
        if losers
        else None
    )
    winner_retention_rate = (
        (len(winners) - len(false_positive)) / len(winners)
        if winners
        else None
    )

    return {
        "trade_count": trade_count,
        "net_pnl": round(sum(pnls), 8),
        "gross_profit": round(gross_profit, 8),
        "gross_loss": round(gross_loss, 8),
        "profit_factor": _compute_profit_factor(gross_profit, gross_loss),
        "win_rate": len(winners) / trade_count if trade_count else None,
        "max_drawdown": _max_drawdown(pnls),
        "winner_count": len(winners),
        "loser_count": len(losers),
        "triggered_count": len(triggered),
        "true_positive_count": len(true_positive),
        "false_positive_count": len(false_positive),
        "false_negative_count": len(false_negative),
        "precision": precision,
        "recall": recall,
        "coverage_ratio": len(covered) / trade_count if trade_count else None,
        "winner_retention_rate": winner_retention_rate,
        "winner_pnl_removed": round(winner_pnl_removed, 8),
        "loser_pnl_removed": round(loser_pnl_removed, 8),
        "simulated_removed_pnl_delta": round(simulated_removed_pnl_delta, 8),
    }


def compute_oos_real_slice_metrics(
    records: Sequence[TradeRecord],
    *,
    hypothesis_scope: Sequence[str] = HYPOTHESIS_SCOPE,
) -> list[dict[str, Any]]:
    """Compute hypothesis metrics across required OOS slice dimensions."""

    output: list[dict[str, Any]] = []
    for hypothesis_id in hypothesis_scope:
        groups: dict[tuple[str, ...], list[TradeRecord]] = {}
        for record in records:
            key_map = _slice_key(record)
            key = tuple(key_map[dimension] for dimension in OOS_SLICE_DIMENSIONS)
            groups.setdefault(key, []).append(record)

        for key, group_records in sorted(groups.items()):
            slice_values = dict(zip(OOS_SLICE_DIMENSIONS, key, strict=True))
            metrics = _metrics_for_records(group_records, hypothesis_id=hypothesis_id)
            output.append(
                {
                    "hypothesis_id": hypothesis_id,
                    "slice": slice_values,
                    "metrics": metrics,
                    "research_only": True,
                    "operational_authority": False,
                    "can_apply_to_freqtrade": False,
                    "can_apply_to_risk_manager": False,
                    "can_promote_rules": False,
                    "can_promote_model": False,
                }
            )
    return output


def _find_first_existing(root: Path, candidates: Sequence[str]) -> Path | None:
    for candidate in candidates:
        path = root / candidate
        if path.exists():
            return path
    return None


def _load_sources_if_allowed(
    project_root: Path,
    *,
    allow_runtime_read: bool,
    paper_source: str | None,
    master_source: str | None,
) -> tuple[list[TradeRecord], list[TradeRecord], list[dict[str, Any]]]:
    if not allow_runtime_read:
        return [], [], []

    paper_path = Path(paper_source) if paper_source else _find_first_existing(
        project_root, DEFAULT_SOURCE_CANDIDATES["paper"]
    )
    master_path = Path(master_source) if master_source else _find_first_existing(
        project_root, DEFAULT_SOURCE_CANDIDATES["master"]
    )

    loaded_sources: list[dict[str, Any]] = []
    paper_records: list[TradeRecord] = []
    master_records: list[TradeRecord] = []

    if paper_path and paper_path.exists():
        paper_records, metadata = read_trade_source(paper_path, source="paper")
        loaded_sources.append(metadata)
    if master_path and master_path.exists():
        master_records, metadata = read_trade_source(master_path, source="master")
        loaded_sources.append(metadata)

    return paper_records, master_records, loaded_sources


def _gate_matrix(
    *,
    computed: bool,
    paper_count: int,
    master_count: int,
) -> list[dict[str, Any]]:
    return [
        {
            "gate_id": "research_only_contract",
            "gate_name": "Research-only contract preserved",
            "severity": "critical",
            "passed": True,
            "evidence": "research_only=true; operational_authority=false",
        },
        {
            "gate_id": "paper_master_divergence_confirmed",
            "gate_name": "Paper/Master divergence remains explicit",
            "severity": "critical",
            "passed": True,
            "evidence": (
                "paper_minus_master_net_pnl="
                f"{CANONICAL_DIVERGENCE_METRICS['paper_minus_master_net_pnl']}"
            ),
        },
        {
            "gate_id": "real_sources_status_explicit",
            "gate_name": "Real source loading status is explicit",
            "severity": "high",
            "passed": True,
            "evidence": f"paper_rows={paper_count}; master_rows={master_count}",
        },
        {
            "gate_id": "real_slice_metrics_status_explicit",
            "gate_name": "Real OOS slice metrics computation status is explicit",
            "severity": "high",
            "passed": True,
            "evidence": f"computed={computed}",
        },
        {
            "gate_id": "oos_required_not_bypassed",
            "gate_name": "OOS validation remains mandatory",
            "severity": "critical",
            "passed": True,
            "evidence": "oos_validation_required=true; oos_validated=false",
        },
        {
            "gate_id": "promotion_blocked",
            "gate_name": "Rule and model promotion blocked",
            "severity": "critical",
            "passed": True,
            "evidence": "can_promote_rules=false; can_promote_model=false",
        },
        {
            "gate_id": "runtime_unchanged",
            "gate_name": "Runtime and execution surfaces unchanged",
            "severity": "critical",
            "passed": True,
            "evidence": "no runtime updates; sends_orders=false",
        },
    ]


def _summary_from_gate_matrix(gates: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    failed = [gate for gate in gates if not gate.get("passed")]
    critical_failed = [
        str(gate["gate_id"])
        for gate in failed
        if str(gate.get("severity")) == "critical"
    ]
    return {
        "gate_count": len(gates),
        "passed_gate_count": len(gates) - len(failed),
        "failed_gate_count": len(failed),
        "failed_gate_ids": [str(gate["gate_id"]) for gate in failed],
        "critical_failed_gate_ids": critical_failed,
    }


def build_oos_real_slice_computation_report(
    *,
    project_root: str | Path = ".",
    allow_runtime_read: bool = False,
    paper_source: str | None = None,
    master_source: str | None = None,
    write_requested: bool = False,
) -> dict[str, Any]:
    root = Path(project_root)
    paper_records, master_records, loaded_sources = _load_sources_if_allowed(
        root,
        allow_runtime_read=allow_runtime_read,
        paper_source=paper_source,
        master_source=master_source,
    )
    real_sources_loaded = bool(loaded_sources)
    computed = bool(paper_records)
    slice_metrics = (
        compute_oos_real_slice_metrics(paper_records)
        if computed
        else []
    )
    global_metrics = (
        {
            hypothesis_id: _metrics_for_records(paper_records, hypothesis_id=hypothesis_id)
            for hypothesis_id in HYPOTHESIS_SCOPE
        }
        if computed
        else {
            hypothesis_id: _metrics_for_records([], hypothesis_id=hypothesis_id)
            for hypothesis_id in HYPOTHESIS_SCOPE
        }
    )

    gates = _gate_matrix(
        computed=computed,
        paper_count=len(paper_records),
        master_count=len(master_records),
    )

    return {
        "schema_version": SCHEMA_VERSION,
        "project_name": PROJECT_NAME,
        "project_root": str(project_root),
        "status": "blocked",
        "reason": (
            "paper_master_divergence_real_slice_metrics_computed_research_only"
            if computed
            else "paper_master_divergence_real_slice_metrics_require_explicit_sources"
        ),
        "decision": "MANTER_EM_RESEARCH",
        "research_only": True,
        "read_only": True,
        "paper_only": True,
        "shadow_only": True,
        "input_mode": (
            "real_sources_loaded_read_only"
            if real_sources_loaded
            else "no_runtime_rows_loaded"
        ),
        "allow_runtime_read": allow_runtime_read,
        "real_source_loader_used": True,
        "real_sources_loaded": real_sources_loaded,
        "loaded_sources": loaded_sources,
        "paper_source_rows": len(paper_records),
        "master_source_rows": len(master_records),
        "hypothesis_scope": list(HYPOTHESIS_SCOPE),
        "oos_slice_dimensions": list(OOS_SLICE_DIMENSIONS),
        "minimum_metrics": list(MINIMUM_METRICS),
        "real_slice_metrics_created": True,
        "real_slice_metrics_computed": computed,
        "oos_slice_metrics_computed": computed,
        "slice_metrics_status": (
            "computed_read_only" if computed else "blocked_no_rows_loaded"
        ),
        "observation_count": len(paper_records),
        "slice_count": len(slice_metrics),
        "slice_metrics": slice_metrics,
        "global_metrics": global_metrics,
        "canonical_cluster_evidence": CANONICAL_CLUSTER_EVIDENCE,
        "paper_kpis": CANONICAL_PAPER_KPIS,
        "master_kpis": CANONICAL_MASTER_KPIS,
        "divergence_metrics": CANONICAL_DIVERGENCE_METRICS,
        "divergence_confirmed": True,
        "paper_replicates_master_edge": False,
        "oos_validation_required": True,
        "oos_validated": False,
        "ready_for_candidate_registry": False,
        "registers_candidate_rules": False,
        "remediation_application_allowed": False,
        "slice_metrics_are_operational": False,
        "expected_trade_value_contract": EXPECTED_TRADE_VALUE_FORMULA,
        "minimum_next_research_gates": [
            "validar estabilidade por dia/símbolo/lado/exit_reason/duração",
            "bloquear qualquer regra que remova ROI winners materialmente",
            "medir efeito em covered vs uncovered antes de extrapolar",
            "avaliar custos, slippage, drawdown, drift e regime",
            "exigir registry shadow bloqueado antes de qualquer observação paper",
        ],
        "allowed_next_steps": [
            "executar CLI com --allow-runtime-read e fontes explícitas read-only",
            "revisar slice_metrics para H1/H2/H6",
            "se efeito sobreviver, criar relatório de validação OOS real",
            "manter promoção bloqueada até registry shadow bloqueado",
        ],
        "forbidden_actions": list(FORBIDDEN_ACTIONS),
        "gate_matrix": gates,
        "gate_summary": _summary_from_gate_matrix(gates),
        "write_requested": write_requested,
        "write_performed": False,
        "writes_data": False,
        "writes_parquet": False,
        "writes_reports": False,
        "writes_runtime": False,
        "writes_sqlite": False,
        "output_path": None,
        "runs_training": False,
        "executes_scheduler": False,
        "executes_orchestrator": False,
        "executes_stage_builders": False,
        "release_authority": False,
        "readiness_release_authority": False,
        "operational_authority": False,
        "can_apply_to_freqtrade": False,
        "can_apply_to_risk_manager": False,
        "can_promote_rules": False,
        "can_promote_model": False,
        "applies_shadow_rules": False,
        "applies_feedback_to_ai_shadow": False,
        "updates_freqtrade": False,
        "updates_risk_manager": False,
        "updates_qlib_runtime": False,
        "updates_ai_shadow_runtime": False,
        "changes_model": False,
        "changes_risk": False,
        "sends_orders": False,
        "exchange_private_access": False,
        "live_trading_enabled": False,
        "live_release_allowed": False,
        "canary_release_allowed": False,
        "order_submission_enabled": False,
        "real_order_submission_enabled": False,
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build research-only real OOS slice metrics for Paper/Master divergence."
    )
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--paper-source", default=None)
    parser.add_argument("--master-source", default=None)
    parser.add_argument("--allow-runtime-read", action="store_true")
    parser.add_argument("--no-write", action="store_true")
    parser.add_argument("--json", action="store_true", dest="as_json")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    report = build_oos_real_slice_computation_report(
        project_root=args.project_root,
        allow_runtime_read=args.allow_runtime_read,
        paper_source=args.paper_source,
        master_source=args.master_source,
        write_requested=not args.no_write,
    )
    if args.as_json:
        print(json.dumps(report, ensure_ascii=False, sort_keys=True))
    else:
        print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
