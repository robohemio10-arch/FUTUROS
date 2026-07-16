"""Filesystem primitives for the guarded two-master transaction."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from .contract import TransitionContract, file_sha256
from .planner import semantic_rows_sha256


@dataclass(frozen=True)
class BackupEvidence:
    directory: Path
    xlsx_path: Path
    parquet_path: Path
    xlsx_sha256: str
    parquet_sha256: str
    xlsx_size: int
    parquet_size: int


@dataclass(frozen=True)
class CandidateEvidence:
    xlsx_path: Path
    parquet_path: Path
    xlsx_row_count: int
    parquet_row_count: int
    xlsx_semantic_sha256: str
    parquet_semantic_sha256: str
    candidate_tail_semantic_sha256: str
    xlsx_size: int
    parquet_size: int


class TransitionLock:
    def __init__(self, path: Path, transition_id: str) -> None:
        self.path = path
        self.transition_id = transition_id
        self.created = False

    def acquire(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        descriptor = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
        payload = {
            "transition_id": self.transition_id,
            "pid": os.getpid(),
            "created_at_utc": datetime.now(UTC).isoformat(),
        }
        try:
            os.write(descriptor, (json.dumps(payload, sort_keys=True) + "\n").encode("utf-8"))
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
        self.created = True

    def release(self) -> None:
        if self.created:
            self.path.unlink(missing_ok=True)
            self.created = False

    def __enter__(self) -> TransitionLock:
        self.acquire()
        return self

    def __exit__(self, _type: object, _value: object, _traceback: object) -> None:
        self.release()


def create_verified_backups(
    *,
    root: Path,
    contract: TransitionContract,
    run_id: str,
) -> BackupEvidence:
    directory = root / "data/backups/bitradex_ocr_legacy_append" / run_id
    directory.mkdir(parents=True, exist_ok=False)
    source_xlsx = root / contract.pre_state.master_xlsx_path
    source_parquet = root / contract.pre_state.master_parquet_path
    xlsx = directory / source_xlsx.name
    parquet = directory / source_parquet.name
    _copy_fsync(source_xlsx, xlsx)
    _copy_fsync(source_parquet, parquet)
    evidence = BackupEvidence(
        directory=directory,
        xlsx_path=xlsx,
        parquet_path=parquet,
        xlsx_sha256=file_sha256(xlsx),
        parquet_sha256=file_sha256(parquet),
        xlsx_size=xlsx.stat().st_size,
        parquet_size=parquet.stat().st_size,
    )
    if evidence.xlsx_sha256 != contract.pre_state.master_xlsx_sha256:
        raise RuntimeError("xlsx_backup_hash_mismatch")
    if evidence.parquet_sha256 != contract.pre_state.master_parquet_sha256:
        raise RuntimeError("parquet_backup_hash_mismatch")
    if evidence.xlsx_size != source_xlsx.stat().st_size or evidence.parquet_size != source_parquet.stat().st_size:
        raise RuntimeError("backup_size_mismatch")
    return evidence


def build_verified_candidates(
    *,
    root: Path,
    contract: TransitionContract,
    candidates: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> CandidateEvidence:
    master_xlsx = root / contract.pre_state.master_xlsx_path
    master_parquet = root / contract.pre_state.master_parquet_path
    parquet_temp = _temporary_peer(master_parquet, ".candidate.parquet")
    xlsx_temp = _temporary_peer(master_xlsx, ".candidate.xlsx")
    try:
        source_parquet_rows = _read_parquet_rows(master_parquet, columns)
        source_xlsx_rows = _read_xlsx_rows(
            master_xlsx,
            columns,
            contract.append_state.xlsx_sheet,
        )
        parquet_rows = _build_parquet_candidate(master_parquet, parquet_temp, candidates, columns)
        xlsx_rows = _build_xlsx_candidate(
            master_xlsx,
            xlsx_temp,
            candidates,
            columns,
            contract.append_state.xlsx_sheet,
        )
        parquet_all = _read_parquet_rows(parquet_temp, columns)
        xlsx_all = _read_xlsx_rows(xlsx_temp, columns, contract.append_state.xlsx_sheet)
        expected = contract.append_state.expected_post_row_count
        if parquet_rows != expected or xlsx_rows != expected:
            raise RuntimeError("candidate_row_count_mismatch")
        prefix_length = contract.pre_state.master_row_count
        if semantic_rows_sha256(parquet_all[:prefix_length], columns) != semantic_rows_sha256(
            source_parquet_rows, columns
        ):
            raise RuntimeError("parquet_candidate_prefix_semantic_mismatch")
        if semantic_rows_sha256(xlsx_all[:prefix_length], columns) != semantic_rows_sha256(
            source_xlsx_rows, columns
        ):
            raise RuntimeError("xlsx_candidate_prefix_semantic_mismatch")
        tail_hash = semantic_rows_sha256(candidates, columns)
        parquet_tail = semantic_rows_sha256(parquet_all[-len(candidates) :], columns)
        xlsx_tail = semantic_rows_sha256(xlsx_all[-len(candidates) :], columns)
        if tail_hash != parquet_tail or tail_hash != xlsx_tail:
            raise RuntimeError("candidate_tail_semantic_mismatch")
        xlsx_semantic_hash = semantic_rows_sha256(xlsx_all, columns)
        parquet_semantic_hash = semantic_rows_sha256(parquet_all, columns)
        if xlsx_semantic_hash != parquet_semantic_hash:
            raise RuntimeError("candidate_cross_format_semantic_mismatch")
        return CandidateEvidence(
            xlsx_path=xlsx_temp,
            parquet_path=parquet_temp,
            xlsx_row_count=xlsx_rows,
            parquet_row_count=parquet_rows,
            xlsx_semantic_sha256=xlsx_semantic_hash,
            parquet_semantic_sha256=parquet_semantic_hash,
            candidate_tail_semantic_sha256=tail_hash,
            xlsx_size=xlsx_temp.stat().st_size,
            parquet_size=parquet_temp.stat().st_size,
        )
    except Exception:
        xlsx_temp.unlink(missing_ok=True)
        parquet_temp.unlink(missing_ok=True)
        raise


def restore_from_backups(
    *, master_xlsx: Path, master_parquet: Path, backup: BackupEvidence
) -> bool:
    xlsx_restore = _temporary_peer(master_xlsx, ".rollback.xlsx")
    parquet_restore = _temporary_peer(master_parquet, ".rollback.parquet")
    try:
        _copy_fsync(backup.xlsx_path, xlsx_restore)
        _copy_fsync(backup.parquet_path, parquet_restore)
        os.replace(parquet_restore, master_parquet)
        os.replace(xlsx_restore, master_xlsx)
        return (
            file_sha256(master_xlsx) == backup.xlsx_sha256
            and file_sha256(master_parquet) == backup.parquet_sha256
        )
    finally:
        xlsx_restore.unlink(missing_ok=True)
        parquet_restore.unlink(missing_ok=True)


def _build_parquet_candidate(
    source: Path,
    destination: Path,
    candidates: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
) -> int:
    master = pq.read_table(source)
    if tuple(master.column_names) != tuple(columns):
        raise RuntimeError("parquet_master_schema_mismatch")
    candidate_table = pa.Table.from_pylist(
        [{column: row.get(column) for column in columns} for row in candidates],
        schema=master.schema,
    )
    for index, column in enumerate(columns):
        source_nonempty = sum(row.get(column) not in (None, "") for row in candidates)
        result_nonnull = candidate_table.column(index).length() - candidate_table.column(index).null_count
        if result_nonnull < source_nonempty:
            raise RuntimeError(f"silent_null_coercion:{column}")
    combined = pa.concat_tables([master, candidate_table])
    pq.write_table(combined, destination)
    _fsync_file(destination)
    reopened = pq.read_table(destination)
    if reopened.schema != master.schema:
        raise RuntimeError("parquet_candidate_schema_drift")
    return reopened.num_rows


def _build_xlsx_candidate(
    source: Path,
    destination: Path,
    candidates: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
    sheet_name: str,
) -> int:
    shutil.copy2(source, destination)
    workbook = load_workbook(destination)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError("xlsx_master_sheet_missing")
        worksheet = workbook[sheet_name]
        header = [cell.value for cell in worksheet[1]][: len(columns)]
        if tuple(header) != tuple(columns):
            raise RuntimeError("xlsx_master_header_mismatch")
        for row in candidates:
            worksheet.append([row.get(column) for column in columns])
        workbook.save(destination)
    finally:
        workbook.close()
    _fsync_file(destination)
    return len(_read_xlsx_rows(destination, columns, sheet_name))


def _read_parquet_rows(path: Path, columns: Sequence[str]) -> list[dict[str, Any]]:
    table = pq.read_table(path, columns=list(columns))
    return [dict(row) for row in table.to_pylist()]


def _read_xlsx_rows(
    path: Path, columns: Sequence[str], sheet_name: str
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError("xlsx_master_sheet_missing")
        iterator = workbook[sheet_name].iter_rows(values_only=True)
        header = tuple(next(iterator, ()))[: len(columns)]
        if header != tuple(columns):
            raise RuntimeError("xlsx_master_header_mismatch")
        return [
            dict(zip(columns, values[: len(columns)], strict=True))
            for values in iterator
            if any(value is not None for value in values[: len(columns)])
        ]
    finally:
        workbook.close()


def _copy_fsync(source: Path, destination: Path) -> None:
    shutil.copy2(source, destination)
    _fsync_file(destination)


def _fsync_file(path: Path) -> None:
    with path.open("r+b") as handle:
        handle.flush()
        os.fsync(handle.fileno())


def _temporary_peer(destination: Path, suffix: str) -> Path:
    descriptor, name = tempfile.mkstemp(prefix=f".{destination.name}.", suffix=suffix, dir=destination.parent)
    os.close(descriptor)
    path = Path(name)
    path.unlink()
    return path
