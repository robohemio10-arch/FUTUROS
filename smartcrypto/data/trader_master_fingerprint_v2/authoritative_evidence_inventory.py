"""Read-only inventory of evidence that may explain legacy Trader Master gaps."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from smartcrypto.security.evidence_bundle_redaction import contains_secret

from .authoritative_sqlite import inspect_sqlite_schema_readonly
from .fingerprint_spec import FINGERPRINT_SPEC_VERSION, HEX_SHA256
from .legacy_lineage_profile import (
    build_source_cohort_profiles,
    profile_legacy_master_row,
)
from .master_adapter import file_sha256, read_trader_master_readonly
from .source_profile import SourceProfileError, load_source_profile


SCHEMA_VERSION = "trader_master_authoritative_evidence_inventory_v2"
DECLARATION_SCHEMA_VERSION = "trader_master_authoritative_evidence_v2"
DEFAULT_MASTER = Path("data/trades/trades_master.parquet")
DEFAULT_JSON_REPORT = Path(
    "data/reports/trader_master_authoritative_evidence_inventory_v2.json"
)
DEFAULT_MARKDOWN_REPORT = Path(
    "data/reports/trader_master_authoritative_evidence_inventory_v2.md"
)
DEFAULT_EVIDENCE_ROOTS = (Path("data"), Path("backups"), Path("docs"), Path("config"))
SOURCE_COHORTS = ("full_ocr_3141", "manual_queue_resolved")
PRIORITY_FIELDS = (
    "account_scope_hash",
    "order_id_namespace",
    "source_trade_id",
    "market_type",
    "contract_type",
    "settlement_currency",
    "quantity_unit",
    "contract_size",
    "gross_pnl",
    "trading_fee",
    "funding_fee",
    "epsilon_abs_fonte",
)
AUTHORITY_CLASSIFICATIONS = (
    "authoritative_and_joinable",
    "authoritative_but_not_joinable",
    "informational_only",
    "conflicting",
    "missing",
)
JOIN_CLASSIFICATIONS = (
    "exact_native_id",
    "exact_source_row_provenance",
    "versioned_deterministic_composite_key",
    "cohort_level_only",
    "not_joinable",
)
BRIDGE_JOIN_CLASSIFICATIONS = frozenset(JOIN_CLASSIFICATIONS[:3])
STRUCTURED_EXTENSIONS = frozenset(
    {".json", ".jsonl", ".csv", ".parquet", ".xlsx", ".sqlite", ".db", ".yaml", ".yml"}
)
TEXT_EXTENSIONS = frozenset({".json", ".jsonl", ".csv", ".md", ".txt", ".yaml", ".yml"})
BINARY_INVENTORY_EXTENSIONS = frozenset({".png", ".jpg", ".jpeg", ".webp", ".pdf", ".zip"})
ALLOWED_EXTENSIONS = STRUCTURED_EXTENSIONS | TEXT_EXTENSIONS | BINARY_INVENTORY_EXTENSIONS
EXCLUDED_DIRECTORY_NAMES = frozenset(
    {".git", ".venv", "venv", "node_modules", "__pycache__", ".pytest_cache", ".mypy_cache", ".ruff_cache", "tmp", "temp"}
)
FORBIDDEN_FILE_NAMES = frozenset(
    {".env", "credentials.json", "credential.json", "secrets.json", "secret.json", "auth.json"}
)
FORBIDDEN_NAME_TOKENS = ("private_key", "api_key", "access_token", "refresh_token")
RELEVANCE_TOKENS = (
    "full_ocr_3141",
    "manual_queue_resolved",
    "trades_master",
    "trader_master",
    "bitradex",
    "ocr",
    "phase5",
    "phase_05",
    "source_profile",
    "import",
    "rebuild",
    "batch",
    "evidence",
    "order",
    "fill",
    "fee",
    "funding",
    "contract",
)
FIELD_COLUMN_ALIASES: dict[str, frozenset[str]] = {
    "account_scope_hash": frozenset({"account_scope_hash"}),
    "order_id_namespace": frozenset({"order_id_namespace"}),
    "source_trade_id": frozenset({"source_trade_id"}),
    "market_type": frozenset({"market_type"}),
    "contract_type": frozenset({"contract_type"}),
    "settlement_currency": frozenset({"settlement_currency"}),
    "quantity_unit": frozenset({"quantity_unit"}),
    "contract_size": frozenset({"contract_size"}),
    "gross_pnl": frozenset({"gross_pnl"}),
    "trading_fee": frozenset(
        {"trading_fee", "fee_open_cost", "fee_close_cost", "fee_open", "fee_close", "taxa_1", "taxa_2"}
    ),
    "funding_fee": frozenset({"funding_fee", "funding_fees"}),
    "epsilon_abs_fonte": frozenset({"epsilon_abs_fonte"}),
}
SOURCE_COHORT_COLUMNS = ("source_cohort", "source_file", "candidate_source")
SOURCE_ROW_COLUMNS = frozenset(
    {"source_row_index", "candidate_source_row_index", "source_line_number", "image_sha256"}
)
NATIVE_ID_COLUMNS = frozenset({"source_trade_id"})
SECRET_SCANNER_PUBLIC_CONTRACT_KEYS = (
    "authoritative_evidence_absence_proven",
)
INSTRUMENT_CONTRACT_FIELDS = frozenset(
    {"market_type", "contract_type", "settlement_currency", "quantity_unit", "contract_size"}
)
FINANCIAL_FIELDS = frozenset({"gross_pnl", "trading_fee", "funding_fee", "epsilon_abs_fonte"})

SAFETY_FLAGS: dict[str, bool] = {
    "fingerprint_generation_allowed": False,
    "bridge_applied": False,
    "import_performed": False,
    "preview_only": True,
    "research_only": True,
    "operational_authority": False,
    "writes_trader_master": False,
    "writes_parquet": False,
    "writes_xlsx": False,
    "writes_csv": False,
    "writes_sqlite": False,
    "writes_runtime": False,
    "changes_fingerprint_spec": False,
    "sends_orders": False,
    "exchange_private_access": False,
}


@dataclass(frozen=True)
class EvidenceArtifact:
    evidence_id: str
    relative_path: str
    artifact_type: str
    sha256: str | None
    size_bytes: int | None
    producer: str | None
    source_cohort: tuple[str, ...]
    batch_id: str | None
    schema_columns: tuple[str, ...]
    candidate_fields: tuple[str, ...]
    authority_classification: str
    join_classification: str
    join_fields: tuple[str, ...]
    row_count: int | None
    integrity_preserved: bool
    sensitive_content_detected: bool
    inspection_status: str
    blockers: tuple[str, ...]
    field_semantics: tuple[str, ...] = ()
    field_value_digests: tuple[tuple[str, str], ...] = ()
    join_diagnostics: tuple[tuple[str, str], ...] = ()
    temp_copy_used: bool = False
    archive_extracted: bool = False
    ocr_executed: bool = False
    sqlite_query_only: bool = False

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["source_cohort"] = list(self.source_cohort)
        payload["schema_columns"] = list(self.schema_columns)
        payload["candidate_fields"] = list(self.candidate_fields)
        payload["join_fields"] = list(self.join_fields)
        payload["blockers"] = list(self.blockers)
        payload["field_semantics"] = list(self.field_semantics)
        payload["field_value_digests"] = dict(self.field_value_digests)
        payload["join_diagnostics"] = dict(self.join_diagnostics)
        return payload


@dataclass(frozen=True)
class DiscoveryResult:
    candidate_paths: tuple[Path, ...]
    blocked_symlinks: tuple[Path, ...]
    evidence_roots_inspected: tuple[str, ...]
    discovered_file_count: int
    ignored_forbidden_file_count: int
    warnings: tuple[str, ...]
    validation_errors: tuple[str, ...]


def build_trader_master_authoritative_evidence_inventory_report(
    *,
    project_root: str | Path,
    trader_master_path: str | Path = DEFAULT_MASTER,
    source_profile_path: str | Path,
    account_scope_hash: str | None,
    authoritative_sqlite_path: str | Path | None = None,
    evidence_roots: Sequence[str | Path] | None = None,
    write_report: bool = False,
    output_json: str | Path = DEFAULT_JSON_REPORT,
    output_markdown: str | Path = DEFAULT_MARKDOWN_REPORT,
    generated_at_utc: str | None = None,
) -> dict[str, Any]:
    """Inventory possible evidence without applying it to the legacy Master."""

    root = Path(project_root).resolve()
    json_path = _resolve(root, output_json)
    markdown_path = _resolve(root, output_markdown)
    report = _base_report(
        root=root,
        trader_master_path=trader_master_path,
        source_profile_path=source_profile_path,
        evidence_roots=evidence_roots,
        write_report=write_report,
        json_path=json_path,
        markdown_path=markdown_path,
        generated_at_utc=generated_at_utc,
    )
    write_errors = _validate_report_paths(root, json_path, markdown_path) if write_report else []
    if write_errors:
        return _blocked(report, "unsafe_report_output_path", write_errors)
    normalized_account_hash = (account_scope_hash or "").strip().casefold()
    if not normalized_account_hash:
        return _blocked(report, "account_scope_hash_missing")
    if HEX_SHA256.fullmatch(normalized_account_hash) is None:
        return _blocked(report, "account_scope_hash_invalid")
    try:
        profile = load_source_profile(_resolve(root, source_profile_path))
    except SourceProfileError as exc:
        return _blocked(report, "source_profile_invalid", str(exc).split(";"))

    master_bundle = read_trader_master_readonly(
        project_root=root,
        trader_master_path=trader_master_path,
    )
    report.update(master_bundle.report)
    if master_bundle.report.get("status") != "ok":
        return _blocked(report, str(master_bundle.report.get("reason", "trader_master_unreadable")))
    if len(master_bundle.source_rows) != int(master_bundle.report.get("trader_master_row_count", -1)):
        return _blocked(report, "trader_master_source_rows_incomplete")

    row_profiles = [
        profile_legacy_master_row(index, row)
        for index, row in enumerate(master_bundle.source_rows)
    ]
    cohort_profiles = build_source_cohort_profiles(master_bundle.source_rows, row_profiles)
    cohorts = _master_source_cohorts(cohort_profiles)
    roots = evidence_roots or DEFAULT_EVIDENCE_ROOTS
    explicit_paths = {
        _resolve(root, source_profile_path),
        _resolve(root, authoritative_sqlite_path or profile.authoritative_sqlite.snapshot_path),
    }
    master_source = _resolve(root, trader_master_path)
    discovery = discover_evidence_candidates(
        project_root=root,
        evidence_roots=roots,
        explicit_paths=explicit_paths,
        excluded_paths={master_source},
    )
    report.update(
        evidence_roots_inspected=list(discovery.evidence_roots_inspected),
        discovered_file_count=discovery.discovered_file_count,
        ignored_forbidden_file_count=discovery.ignored_forbidden_file_count,
    )
    if discovery.validation_errors:
        return _blocked(report, "unsafe_evidence_root", discovery.validation_errors)

    artifacts = [
        _blocked_symlink_artifact(path, root) for path in discovery.blocked_symlinks
    ]
    for path in discovery.candidate_paths:
        artifacts.append(
            inspect_evidence_artifact(
                project_root=root,
                path=path,
                expected_account_scope_hash=normalized_account_hash,
            )
        )
    artifacts.sort(key=lambda artifact: artifact.relative_path)
    evidence_by_field = build_evidence_by_cohort_and_field(cohorts, artifacts)
    classification_counts = Counter(
        str(item["selected_classification"]) for item in evidence_by_field
    )
    resolvable_by_cohort = {
        cohort: sum(
            bool(item["bridge_field_resolvable_without_fabrication"])
            for item in evidence_by_field
            if item["source_cohort"] == cohort
        )
        for cohort in cohorts
    }
    all_resolvable = {
        cohort: resolvable_by_cohort[cohort] == len(PRIORITY_FIELDS) for cohort in cohorts
    }
    decision = evidence_decision(classification_counts, all_resolvable)
    financial_by_cohort = build_financial_evidence_summary(evidence_by_field, cohorts)
    artifact_candidate_count = len(artifacts)
    artifact_inspected_count = sum(
        artifact.inspection_status == "inspected" for artifact in artifacts
    )
    artifact_blocked_count = sum(bool(artifact.blockers) for artifact in artifacts)
    coverage = build_inventory_coverage(
        artifact_candidate_count=artifact_candidate_count,
        artifact_inspected_count=artifact_inspected_count,
        artifact_blocked_count=artifact_blocked_count,
        decision=decision,
    )
    if not coverage["inventory_accounting_consistent"]:
        report.update(
            artifact_candidate_count=artifact_candidate_count,
            artifact_inspected_count=artifact_inspected_count,
            artifact_blocked_count=artifact_blocked_count,
            **coverage,
        )
        return _blocked(report, "inventory_accounting_inconsistent")
    report.update(
        status="ok",
        reason=_inventory_reason(decision, bool(coverage["inventory_coverage_complete"])),
        decision=decision,
        trader_master_sha256=master_bundle.report.get("trader_master_sha256_before"),
        source_cohorts=list(cohorts),
        source_cohort_profiles=cohort_profiles,
        artifact_candidate_count=artifact_candidate_count,
        artifact_inspected_count=artifact_inspected_count,
        artifact_blocked_count=artifact_blocked_count,
        **coverage,
        evidence_inventory=[artifact.to_dict() for artifact in artifacts],
        evidence_by_cohort_and_field=evidence_by_field,
        authoritative_and_joinable_count=int(
            classification_counts["authoritative_and_joinable"]
        ),
        authoritative_but_not_joinable_count=int(
            classification_counts["authoritative_but_not_joinable"]
        ),
        informational_only_count=int(classification_counts["informational_only"]),
        conflicting_count=int(classification_counts["conflicting"]),
        missing_field_count=int(classification_counts["missing"]),
        full_ocr_3141_resolvable_field_count=resolvable_by_cohort.get("full_ocr_3141", 0),
        manual_queue_resolved_resolvable_field_count=resolvable_by_cohort.get(
            "manual_queue_resolved", 0
        ),
        all_required_fields_resolvable_by_cohort=all_resolvable,
        bridge_design_preconditions_satisfied=bool(all_resolvable) and all(
            all_resolvable.values()
        ),
        financial_evidence_by_cohort=financial_by_cohort,
        recommended_next_action=_recommended_action(decision),
        account_scope_hash_present=True,
        account_scope_hash_valid=True,
        account_scope_original_identifier_persisted=False,
        validation_errors=[],
        blockers=[],
        warnings=sorted(set(discovery.warnings) | set(_inventory_warnings(artifacts))),
        import_eligible_true_count=0,
        **SAFETY_FLAGS,
        safety_flags=dict(SAFETY_FLAGS),
    )
    return _maybe_write(report, write_report, json_path, markdown_path)


def discover_evidence_candidates(
    *,
    project_root: Path,
    evidence_roots: Sequence[str | Path],
    explicit_paths: Iterable[Path] = (),
    excluded_paths: Iterable[Path] = (),
) -> DiscoveryResult:
    root = project_root.resolve()
    excluded = {path.resolve() for path in excluded_paths}
    candidates: set[Path] = set()
    blocked_symlinks: set[Path] = set()
    inspected_roots: list[str] = []
    warnings: list[str] = []
    validation_errors: list[str] = []
    discovered_count = 0
    forbidden_count = 0
    for requested in evidence_roots:
        path = _resolve(root, requested)
        try:
            path.relative_to(root)
        except ValueError:
            validation_errors.append("evidence_root_outside_project_root")
            continue
        if path.is_symlink():
            validation_errors.append("evidence_root_symlink_forbidden")
            continue
        if not path.exists():
            warnings.append(f"evidence_root_missing:{_display_path(path, root)}")
            continue
        if not path.is_dir():
            validation_errors.append("evidence_root_not_directory")
            continue
        inspected_roots.append(_display_path(path, root))
        for current, directories, files in os.walk(path, followlinks=False):
            current_path = Path(current)
            retained_directories: list[str] = []
            for directory in sorted(directories):
                child = current_path / directory
                if directory.casefold() in EXCLUDED_DIRECTORY_NAMES:
                    continue
                if child.is_symlink():
                    blocked_symlinks.add(child)
                    continue
                retained_directories.append(directory)
            directories[:] = retained_directories
            for filename in sorted(files):
                candidate = current_path / filename
                if _forbidden_filename(filename):
                    forbidden_count += 1
                    continue
                if candidate.is_symlink():
                    blocked_symlinks.add(candidate)
                    continue
                if candidate.suffix.casefold() not in ALLOWED_EXTENSIONS:
                    continue
                discovered_count += 1
                if candidate.resolve() in excluded:
                    continue
                if _path_is_relevant_candidate(candidate, root):
                    candidates.add(candidate.resolve())
    for explicit in explicit_paths:
        path = explicit.resolve()
        try:
            path.relative_to(root)
        except ValueError:
            validation_errors.append("explicit_evidence_outside_project_root")
            continue
        if path in excluded or not path.exists():
            continue
        if path.is_symlink():
            blocked_symlinks.add(path)
        elif path.is_file() and not _forbidden_filename(path.name):
            candidates.add(path)
    return DiscoveryResult(
        candidate_paths=tuple(sorted(candidates, key=lambda item: item.as_posix())),
        blocked_symlinks=tuple(sorted(blocked_symlinks, key=lambda item: item.as_posix())),
        evidence_roots_inspected=tuple(sorted(set(inspected_roots))),
        discovered_file_count=discovered_count,
        ignored_forbidden_file_count=forbidden_count,
        warnings=tuple(sorted(set(warnings))),
        validation_errors=tuple(sorted(set(validation_errors))),
    )


def inspect_evidence_artifact(
    *,
    project_root: Path,
    path: Path,
    expected_account_scope_hash: str | None = None,
) -> EvidenceArtifact:
    root = project_root.resolve()
    relative = _display_path(path, root)
    path_error = _validate_artifact_path(root, path)
    if path_error is not None:
        return _empty_artifact(relative, path.suffix.casefold(), path_error)
    before_hash = file_sha256(path)
    before_size = path.stat().st_size
    extension = path.suffix.casefold()
    if extension in {".sqlite", ".db"}:
        metadata = _inspect_sqlite_artifact(root, path)
        after_hash = file_sha256(path)
        after_size = path.stat().st_size
        return _artifact_from_metadata(
            relative=relative,
            extension=extension,
            sha256=before_hash,
            size_bytes=before_size,
            metadata=metadata,
            integrity_preserved=(before_hash == after_hash and before_size == after_size),
            temp_copy_used=bool(metadata.get("temp_copy_used")),
            sqlite_query_only=bool(metadata.get("sqlite_query_only")),
        )
    try:
        with TemporaryDirectory(prefix="trader-master-evidence-") as temporary:
            copied = Path(temporary) / path.name
            shutil.copy2(path, copied)
            metadata = _inspect_copied_artifact(
                copied,
                relative,
                extension,
                expected_account_scope_hash=expected_account_scope_hash,
            )
    except (OSError, ValueError, json.JSONDecodeError, UnicodeError, csv.Error) as exc:
        metadata = {
            "inspection_status": "blocked",
            "blockers": [f"artifact_unreadable:{type(exc).__name__}"],
        }
    after_hash = file_sha256(path)
    after_size = path.stat().st_size
    return _artifact_from_metadata(
        relative=relative,
        extension=extension,
        sha256=before_hash,
        size_bytes=before_size,
        metadata=metadata,
        integrity_preserved=(before_hash == after_hash and before_size == after_size),
        temp_copy_used=True,
        sqlite_query_only=False,
    )


def build_evidence_by_cohort_and_field(
    cohorts: Sequence[str],
    artifacts: Sequence[EvidenceArtifact],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for cohort in cohorts:
        for field in PRIORITY_FIELDS:
            candidates = [
                artifact
                for artifact in artifacts
                if cohort in artifact.source_cohort and field in artifact.candidate_fields
            ]
            authoritative = [
                artifact
                for artifact in candidates
                if artifact.authority_classification
                in {"authoritative_and_joinable", "authoritative_but_not_joinable"}
            ]
            joinable = [
                artifact
                for artifact in authoritative
                if artifact.authority_classification == "authoritative_and_joinable"
                and artifact.join_classification in BRIDGE_JOIN_CLASSIFICATIONS
                and not artifact.blockers
            ]
            conflict = _joinable_evidence_conflicts(joinable, field)
            if conflict:
                classification = "conflicting"
                selected = None
                join_classification = "not_joinable"
                missing_reason = "authoritative_sources_conflict"
            elif joinable:
                classification = "authoritative_and_joinable"
                selected = sorted(joinable, key=lambda item: item.evidence_id)[0]
                join_classification = selected.join_classification
                missing_reason = None
            elif authoritative:
                classification = "authoritative_but_not_joinable"
                selected = None
                join_classification = "not_joinable"
                missing_reason = "authoritative_evidence_lacks_exact_join"
            elif candidates:
                classification = "informational_only"
                selected = None
                join_classification = "cohort_level_only"
                missing_reason = "no_explicit_authoritative_joinable_contract"
            else:
                classification = "missing"
                selected = None
                join_classification = "not_joinable"
                missing_reason = "no_candidate_evidence_for_cohort_field"
            resolvable = classification == "authoritative_and_joinable"
            results.append(
                {
                    "source_cohort": cohort,
                    "canonical_field": field,
                    "evidence_candidates": sorted(
                        artifact.evidence_id for artifact in candidates
                    ),
                    "authoritative_candidate_count": len(authoritative),
                    "joinable_candidate_count": len(joinable),
                    "conflicting_candidate_count": len(joinable) if conflict else 0,
                    "selected_classification": classification,
                    "selected_evidence_id": selected.evidence_id if selected else None,
                    "join_classification": join_classification,
                    "missing_evidence_reason": missing_reason,
                    "bridge_field_resolvable": resolvable,
                    "bridge_field_resolvable_without_fabrication": resolvable,
                    "external_manual_attestation_required": bool(
                        field == "account_scope_hash" and not resolvable
                    ),
                }
            )
    return results


def build_financial_evidence_summary(
    evidence_by_field: Sequence[Mapping[str, Any]],
    cohorts: Sequence[str],
) -> dict[str, str]:
    result: dict[str, str] = {}
    financial_fields = {"gross_pnl", "trading_fee", "funding_fee"}
    for cohort in cohorts:
        rows = [
            item
            for item in evidence_by_field
            if item["source_cohort"] == cohort
            and item["canonical_field"] in financial_fields
        ]
        classifications = {str(item["selected_classification"]) for item in rows}
        resolved = sum(bool(item["bridge_field_resolvable"]) for item in rows)
        if "conflicting" in classifications:
            status = "financial_evidence_conflicting"
        elif resolved == len(financial_fields):
            status = "full_financial_evidence_joinable"
        elif resolved:
            status = "partial_financial_evidence_joinable"
        elif classifications & {"authoritative_but_not_joinable", "informational_only"}:
            status = "financial_evidence_not_joinable"
        else:
            status = "financial_evidence_missing"
        result[cohort] = status
    return result


def evidence_decision(
    classification_counts: Mapping[str, int],
    all_resolvable: Mapping[str, bool],
) -> str:
    if int(classification_counts.get("conflicting", 0)):
        return "CONFLICTING_AUTHORITATIVE_EVIDENCE"
    if all_resolvable and all(all_resolvable.values()):
        return "AUTHORITATIVE_EVIDENCE_COMPLETE_AND_JOINABLE"
    if int(classification_counts.get("authoritative_and_joinable", 0)):
        return "PARTIAL_AUTHORITATIVE_EVIDENCE_FOUND"
    if int(classification_counts.get("authoritative_but_not_joinable", 0)):
        return "AUTHORITATIVE_EVIDENCE_NOT_JOINABLE"
    return "NO_AUTHORITATIVE_EVIDENCE_FOUND"


def build_inventory_coverage(
    *,
    artifact_candidate_count: int,
    artifact_inspected_count: int,
    artifact_blocked_count: int,
    decision: str,
) -> dict[str, bool | int | str]:
    """Qualify decision scope without hiding impossible artifact accounting."""

    artifact_uninspected_count = max(
        0,
        artifact_candidate_count - artifact_inspected_count,
    )
    inventory_accounting_consistent = (
        artifact_candidate_count >= artifact_inspected_count
        and artifact_uninspected_count >= artifact_blocked_count
    )
    inventory_coverage_complete = (
        inventory_accounting_consistent
        and artifact_uninspected_count == 0
        and artifact_blocked_count == 0
    )
    authoritative_evidence_absence_proven = (
        decision == "NO_AUTHORITATIVE_EVIDENCE_FOUND"
        and inventory_coverage_complete
    )
    blocked_artifacts_may_contain_unassessed_evidence = artifact_uninspected_count > 0
    decision_scope = (
        "complete_candidate_inventory"
        if inventory_coverage_complete
        else "safely_inspected_artifacts_only"
    )
    return {
        "artifact_uninspected_count": artifact_uninspected_count,
        "inventory_accounting_consistent": inventory_accounting_consistent,
        "inventory_coverage_complete": inventory_coverage_complete,
        "authoritative_evidence_absence_proven": authoritative_evidence_absence_proven,
        "blocked_artifacts_may_contain_unassessed_evidence": (
            blocked_artifacts_may_contain_unassessed_evidence
        ),
        "decision_scope": decision_scope,
    }


def _inventory_reason(decision: str, inventory_coverage_complete: bool) -> str:
    if decision != "NO_AUTHORITATIVE_EVIDENCE_FOUND":
        return "authoritative_evidence_inventory_completed"
    if inventory_coverage_complete:
        return "no_authoritative_evidence_found_after_complete_candidate_inventory"
    return "no_authoritative_evidence_found_in_safely_inspected_artifacts"


def render_markdown(report: Mapping[str, Any]) -> str:
    return "\n".join(
        [
            "# Trader Master Authoritative Evidence Inventory V2",
            "",
            f"- Status: `{report.get('status')}`",
            f"- Decision: `{report.get('decision')}`",
            f"- Master rows: `{report.get('trader_master_row_count')}`",
            f"- Candidate artifacts: `{report.get('artifact_candidate_count')}`",
            f"- Inspected artifacts: `{report.get('artifact_inspected_count')}`",
            f"- Uninspected artifacts: `{report.get('artifact_uninspected_count')}`",
            f"- Blocked artifacts: `{report.get('artifact_blocked_count')}`",
            f"- Inventory coverage complete: `{report.get('inventory_coverage_complete')}`",
            f"- Authoritative absence proven: `{report.get('authoritative_evidence_absence_proven')}`",
            f"- Decision scope: `{report.get('decision_scope')}`",
            f"- Authoritative and joinable field/cohort pairs: `{report.get('authoritative_and_joinable_count')}`",
            f"- Missing field/cohort pairs: `{report.get('missing_field_count')}`",
            f"- Bridge preconditions satisfied: `{report.get('bridge_design_preconditions_satisfied')}`",
            "",
            "## Boundary",
            "",
            "This inventory never applies evidence, creates Fingerprint V2, imports trades, or changes the Trader Master.",
            "",
            f"Recommended next action: `{report.get('recommended_next_action')}`",
            "",
        ]
    )


def _inspect_copied_artifact(
    copied: Path,
    relative_path: str,
    extension: str,
    *,
    expected_account_scope_hash: str | None,
) -> dict[str, Any]:
    if extension in BINARY_INVENTORY_EXTENSIONS:
        return {
            "inspection_status": "inspected",
            "artifact_type": _artifact_type(extension),
            "schema_columns": [],
            "candidate_fields": [],
            "source_cohorts": _cohorts_from_text(relative_path),
            "authority_classification": "informational_only",
            "join_classification": "cohort_level_only"
            if _cohorts_from_text(relative_path)
            else "not_joinable",
            "join_fields": [],
            "blockers": [],
            "archive_extracted": False,
            "ocr_executed": False,
        }
    if extension in TEXT_EXTENSIONS:
        text = copied.read_text(encoding="utf-8-sig")
        if _contains_sensitive_material(text):
            return {
                "inspection_status": "blocked",
                "artifact_type": _artifact_type(extension),
                "sensitive_content_detected": True,
                "blockers": ["sensitive_content_detected"],
            }
        if extension == ".json":
            return _inspect_json(
                json.loads(text),
                extension,
                expected_account_scope_hash=expected_account_scope_hash,
            )
        if extension == ".jsonl":
            return _inspect_jsonl(text)
        if extension == ".csv":
            return _inspect_csv(copied)
        return _inspect_document_text(text, extension)
    if extension == ".parquet":
        return _inspect_parquet(copied)
    if extension == ".xlsx":
        return _inspect_xlsx(copied)
    raise ValueError("unsupported_evidence_extension")


def _contains_sensitive_material(text: str) -> bool:
    sanitized = text
    for field in SECRET_SCANNER_PUBLIC_CONTRACT_KEYS:
        sanitized = sanitized.replace(field, "public_evidence_contract_field")
    return contains_secret(sanitized)


def _inspect_json(
    payload: Any,
    extension: str,
    *,
    expected_account_scope_hash: str | None,
) -> dict[str, Any]:
    keys = _collect_mapping_keys(payload)
    declaration = payload if isinstance(payload, Mapping) else {}
    row_count = len(payload) if isinstance(payload, list) else 1
    metadata = _metadata_from_columns(keys)
    metadata.update(
        artifact_type=_artifact_type(extension),
        row_count=row_count,
        producer=_first_text(declaration, ("producer", "producer_module", "generated_by_script")),
        batch_id=_first_text(declaration, ("batch_id", "ingestion_run_id")),
        source_cohorts=_cohorts_from_payload(payload),
    )
    metadata.update(
        _declaration_metadata(
            declaration,
            metadata["candidate_fields"],
            expected_account_scope_hash=expected_account_scope_hash,
        )
    )
    return metadata


def _inspect_jsonl(text: str) -> dict[str, Any]:
    keys: set[str] = set()
    cohorts: set[str] = set()
    row_count = 0
    for line in text.splitlines():
        if not line.strip():
            continue
        payload = json.loads(line)
        row_count += 1
        keys.update(_collect_mapping_keys(payload))
        cohorts.update(_cohorts_from_payload(payload))
    metadata = _metadata_from_columns(keys)
    metadata.update(
        artifact_type="jsonl",
        row_count=row_count,
        source_cohorts=sorted(cohorts),
    )
    return metadata


def _inspect_csv(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = [str(column) for column in (reader.fieldnames or [])]
        return _inspect_tabular_rows(columns, reader, artifact_type="csv")


def _inspect_parquet(path: Path) -> dict[str, Any]:
    import pyarrow.parquet as parquet

    parquet_file = parquet.ParquetFile(path)
    columns = [str(column) for column in parquet_file.schema.names]
    selected = [
        column
        for column in columns
        if column in set(SOURCE_COHORT_COLUMNS) | SOURCE_ROW_COLUMNS | NATIVE_ID_COLUMNS | {"order_id"}
    ]
    rows: Iterable[Mapping[str, Any]]
    if selected:
        rows = parquet_file.read(columns=selected).to_pylist()
    else:
        rows = ()
    metadata = _inspect_tabular_rows(columns, rows, artifact_type="parquet")
    metadata["row_count"] = int(parquet_file.metadata.num_rows)
    return metadata


def _inspect_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, ())
        columns = [str(value) if value is not None else "" for value in header]
        rows = (
            {column: value for column, value in zip(columns, values, strict=False) if column}
            for values in iterator
        )
        return _inspect_tabular_rows(columns, rows, artifact_type="xlsx")
    finally:
        workbook.close()


def _inspect_tabular_rows(
    columns: Sequence[str],
    rows: Iterable[Mapping[str, Any]],
    *,
    artifact_type: str,
) -> dict[str, Any]:
    cohorts: set[str] = set()
    native_values: dict[str, list[str]] = defaultdict(list)
    row_count = 0
    for row in rows:
        row_count += 1
        for column in SOURCE_COHORT_COLUMNS:
            value = row.get(column)
            if value is not None:
                cohorts.update(_cohorts_from_text(str(value)))
        for column in NATIVE_ID_COLUMNS | {"order_id"}:
            value = row.get(column)
            if value is not None and str(value).strip():
                native_values[column].append(str(value).strip())
    metadata = _metadata_from_columns(columns)
    diagnostics = {
        column: json.dumps(
            {
                "present_count": len(values),
                "distinct_count": len(set(values)),
                "unique_when_present": len(values) == len(set(values)),
            },
            sort_keys=True,
        )
        for column, values in native_values.items()
    }
    metadata.update(
        artifact_type=artifact_type,
        row_count=row_count,
        source_cohorts=sorted(cohorts),
        join_diagnostics=diagnostics,
    )
    return metadata


def _inspect_document_text(text: str, extension: str) -> dict[str, Any]:
    candidate_fields = [field for field in PRIORITY_FIELDS if field in text]
    return {
        "inspection_status": "inspected",
        "artifact_type": _artifact_type(extension),
        "schema_columns": [],
        "candidate_fields": candidate_fields,
        "source_cohorts": _cohorts_from_text(text),
        "authority_classification": "informational_only",
        "join_classification": "cohort_level_only"
        if _cohorts_from_text(text)
        else "not_joinable",
        "join_fields": [],
        "row_count": None,
        "blockers": [],
    }


def _inspect_sqlite_artifact(root: Path, path: Path) -> dict[str, Any]:
    inspected = inspect_sqlite_schema_readonly(project_root=root, snapshot_path=path)
    schemas = inspected.get("table_schemas", {})
    if not isinstance(schemas, Mapping):
        schemas = {}
    columns = sorted(
        {
            str(column)
            for values in schemas.values()
            if isinstance(values, Sequence)
            for column in values
        }
    )
    metadata = _metadata_from_columns(columns)
    metadata.update(
        artifact_type="sqlite",
        row_count=sum(int(value) for value in inspected.get("table_row_counts", {}).values()),
        source_cohorts=[],
        temp_copy_used=bool(inspected.get("snapshot_temp_copy_used")),
        sqlite_query_only=bool(inspected.get("snapshot_query_only")),
        inspection_status="inspected" if inspected.get("status") == "ok" else "blocked",
        blockers=list(inspected.get("validation_errors", [])),
        join_diagnostics={
            table: json.dumps(value, sort_keys=True)
            for table, value in _mapping_items(
                inspected.get("join_column_diagnostics")
            )
        },
    )
    return metadata


def _metadata_from_columns(columns: Iterable[str]) -> dict[str, Any]:
    normalized = {str(column).strip() for column in columns if str(column).strip()}
    candidate_fields = sorted(
        field
        for field, aliases in FIELD_COLUMN_ALIASES.items()
        if normalized & aliases
    )
    join_fields: list[str] = []
    if normalized & NATIVE_ID_COLUMNS:
        join_classification = "exact_native_id"
        join_fields = sorted(normalized & NATIVE_ID_COLUMNS)
    elif normalized & SOURCE_ROW_COLUMNS and normalized & {"source_file", "source_cohort"}:
        join_classification = "exact_source_row_provenance"
        join_fields = sorted(
            (normalized & SOURCE_ROW_COLUMNS) | (normalized & {"source_file", "source_cohort"})
        )
    else:
        join_classification = "not_joinable"
    return {
        "inspection_status": "inspected",
        "schema_columns": sorted(normalized),
        "candidate_fields": candidate_fields,
        "authority_classification": "informational_only",
        "join_classification": join_classification,
        "join_fields": join_fields,
        "blockers": [],
        "field_semantics": [],
        "field_value_digests": {},
        "join_diagnostics": {},
    }


def _declaration_metadata(
    declaration: Mapping[str, Any],
    discovered_candidate_fields: Sequence[str],
    *,
    expected_account_scope_hash: str | None,
) -> dict[str, Any]:
    if declaration.get("schema_version") != DECLARATION_SCHEMA_VERSION:
        return {}
    semantics = declaration.get("field_semantics")
    semantics_fields = (
        sorted(str(field) for field in semantics if field in PRIORITY_FIELDS)
        if isinstance(semantics, Mapping)
        else []
    )
    candidate_fields = sorted(set(discovered_candidate_fields) | set(semantics_fields))
    producer = _first_text(declaration, ("producer", "producer_module"))
    authority_declared = (
        str(declaration.get("provenance_classification", "")).casefold()
        == "authoritative"
    )
    join_contract = declaration.get("join_contract")
    join_classification = "not_joinable"
    join_fields: list[str] = []
    if isinstance(join_contract, Mapping):
        requested = str(join_contract.get("classification", "not_joinable"))
        if requested in JOIN_CLASSIFICATIONS:
            join_classification = requested
        fields = join_contract.get("fields")
        if isinstance(fields, Sequence) and not isinstance(fields, (str, bytes)):
            join_fields = sorted(str(field) for field in fields if str(field).strip())
    complete_semantics = bool(candidate_fields) and set(candidate_fields) <= set(semantics_fields)
    exact_join = (
        join_classification in BRIDGE_JOIN_CLASSIFICATIONS
        and bool(join_fields)
        and isinstance(join_contract, Mapping)
        and join_contract.get("deterministic_per_row") is True
        and join_contract.get("uniqueness_verified") is True
        and join_contract.get("fuzzy_matching") is False
    )
    blockers: list[str] = []
    if not producer:
        blockers.append("authoritative_producer_missing")
    if not complete_semantics:
        blockers.append("field_semantics_incomplete")
    if isinstance(join_contract, Mapping) and not exact_join:
        blockers.append("deterministic_exact_join_not_verified")
    if set(candidate_fields) & INSTRUMENT_CONTRACT_FIELDS and not _valid_instrument_scope(
        declaration.get("instrument_scope")
    ):
        blockers.append("versioned_instrument_scope_missing")
    if "account_scope_hash" in candidate_fields and not _valid_account_attestation(
        declaration.get("account_scope_attestation"),
        expected_account_scope_hash=expected_account_scope_hash,
    ):
        blockers.append("sanitized_account_scope_attestation_missing")
    if set(candidate_fields) & FINANCIAL_FIELDS and not _valid_financial_provenance(
        declaration.get("financial_provenance"),
        set(candidate_fields) & FINANCIAL_FIELDS,
    ):
        blockers.append("financial_source_columns_incomplete")
    if authority_declared and producer and complete_semantics and exact_join:
        authority = "authoritative_and_joinable"
    elif authority_declared and producer and complete_semantics:
        authority = "authoritative_but_not_joinable"
    else:
        authority = "informational_only"
    digests = declaration.get("field_value_digests")
    field_value_digests = {
        str(field): str(value)
        for field, value in _mapping_items(digests)
        if field in PRIORITY_FIELDS
        and isinstance(value, str)
        and HEX_SHA256.fullmatch(value.casefold())
    }
    return {
        "producer": producer,
        "batch_id": _first_text(declaration, ("batch_id",)),
        "source_cohorts": _declared_cohorts(declaration.get("source_cohort")),
        "candidate_fields": candidate_fields,
        "field_semantics": semantics_fields,
        "field_value_digests": field_value_digests,
        "authority_classification": authority,
        "join_classification": join_classification,
        "join_fields": join_fields,
        "blockers": blockers,
    }


def _valid_instrument_scope(value: Any) -> bool:
    if not isinstance(value, Mapping):
        return False
    symbols = value.get("symbols")
    return bool(
        _optional_text(value.get("schema_version"))
        and _optional_text(value.get("market_type"))
        and _optional_text(value.get("valid_from_utc"))
        and _optional_text(value.get("valid_to_utc"))
        and isinstance(symbols, Sequence)
        and not isinstance(symbols, (str, bytes))
        and any(_optional_text(symbol) for symbol in symbols)
    )


def _valid_account_attestation(
    value: Any,
    *,
    expected_account_scope_hash: str | None,
) -> bool:
    if not isinstance(value, Mapping):
        return False
    account_hash = str(value.get("account_scope_hash", "")).casefold()
    return bool(
        HEX_SHA256.fullmatch(account_hash)
        and (
            expected_account_scope_hash is None
            or account_hash == expected_account_scope_hash.casefold()
        )
        and _optional_text(value.get("provenance"))
        and value.get("original_identifier_included") is False
    )


def _valid_financial_provenance(value: Any, fields: set[str]) -> bool:
    if not isinstance(value, Mapping):
        return False
    source_columns = value.get("source_columns")
    formulas = value.get("formulas")
    if not isinstance(source_columns, Mapping) or not isinstance(formulas, Mapping):
        return False
    return all(
        _optional_text(source_columns.get(field)) and _optional_text(formulas.get(field))
        for field in fields
    )


def _artifact_from_metadata(
    *,
    relative: str,
    extension: str,
    sha256: str,
    size_bytes: int,
    metadata: Mapping[str, Any],
    integrity_preserved: bool,
    temp_copy_used: bool,
    sqlite_query_only: bool,
) -> EvidenceArtifact:
    blockers = list(metadata.get("blockers", []))
    if not integrity_preserved:
        blockers.append("artifact_changed_during_inspection")
    sensitive = bool(metadata.get("sensitive_content_detected"))
    if sensitive and "sensitive_content_detected" not in blockers:
        blockers.append("sensitive_content_detected")
    classification = str(metadata.get("authority_classification", "informational_only"))
    if classification not in AUTHORITY_CLASSIFICATIONS:
        classification = "informational_only"
        blockers.append("invalid_authority_classification")
    join_classification = str(metadata.get("join_classification", "not_joinable"))
    if join_classification not in JOIN_CLASSIFICATIONS:
        join_classification = "not_joinable"
        blockers.append("invalid_join_classification")
    if blockers and classification == "authoritative_and_joinable":
        classification = "informational_only"
    evidence_id = "evidence-" + hashlib.sha256(
        f"{relative}\0{sha256}".encode("utf-8")
    ).hexdigest()[:20]
    value_digests = dict(_mapping_items(metadata.get("field_value_digests")))
    join_diagnostics = dict(_mapping_items(metadata.get("join_diagnostics")))
    return EvidenceArtifact(
        evidence_id=evidence_id,
        relative_path=relative,
        artifact_type=str(metadata.get("artifact_type", _artifact_type(extension))),
        sha256=sha256,
        size_bytes=size_bytes,
        producer=_optional_text(metadata.get("producer")),
        source_cohort=tuple(sorted(set(metadata.get("source_cohorts", [])))),
        batch_id=_optional_text(metadata.get("batch_id")),
        schema_columns=tuple(sorted(set(metadata.get("schema_columns", [])))),
        candidate_fields=tuple(sorted(set(metadata.get("candidate_fields", [])))),
        authority_classification=classification,
        join_classification=join_classification,
        join_fields=tuple(sorted(set(metadata.get("join_fields", [])))),
        row_count=_optional_int(metadata.get("row_count")),
        integrity_preserved=integrity_preserved,
        sensitive_content_detected=sensitive,
        inspection_status=(
            "blocked" if blockers else str(metadata.get("inspection_status", "inspected"))
        ),
        blockers=tuple(sorted(set(str(blocker) for blocker in blockers))),
        field_semantics=tuple(sorted(set(metadata.get("field_semantics", [])))),
        field_value_digests=tuple(
            sorted((str(key), str(value)) for key, value in value_digests.items())
        ),
        join_diagnostics=tuple(
            sorted((str(key), str(value)) for key, value in join_diagnostics.items())
        ),
        temp_copy_used=temp_copy_used,
        archive_extracted=False,
        ocr_executed=False,
        sqlite_query_only=sqlite_query_only,
    )


def _mapping_items(value: Any) -> Iterable[tuple[Any, Any]]:
    return value.items() if isinstance(value, Mapping) else ()


def _blocked_symlink_artifact(path: Path, root: Path) -> EvidenceArtifact:
    return _empty_artifact(
        _display_path_without_resolve(path, root),
        path.suffix.casefold(),
        "artifact_symlink_forbidden",
    )


def _empty_artifact(relative: str, extension: str, blocker: str) -> EvidenceArtifact:
    evidence_id = "evidence-" + hashlib.sha256(relative.encode("utf-8")).hexdigest()[:20]
    return EvidenceArtifact(
        evidence_id=evidence_id,
        relative_path=relative,
        artifact_type=_artifact_type(extension),
        sha256=None,
        size_bytes=None,
        producer=None,
        source_cohort=(),
        batch_id=None,
        schema_columns=(),
        candidate_fields=(),
        authority_classification="informational_only",
        join_classification="not_joinable",
        join_fields=(),
        row_count=None,
        integrity_preserved=False,
        sensitive_content_detected=False,
        inspection_status="blocked",
        blockers=(blocker,),
    )


def _joinable_evidence_conflicts(
    artifacts: Sequence[EvidenceArtifact],
    field: str,
) -> bool:
    if len(artifacts) < 2:
        return False
    digests: list[str] = []
    for artifact in artifacts:
        value = dict(artifact.field_value_digests).get(field)
        if value is None:
            return True
        digests.append(value)
    return len(set(digests)) > 1


def _master_source_cohorts(
    profiles: Sequence[Mapping[str, Any]],
) -> tuple[str, ...]:
    observed = {
        str(profile.get("cohort_values", {}).get("source_file"))
        for profile in profiles
        if str(profile.get("cohort_values", {}).get("source_file")) in SOURCE_COHORTS
    }
    return tuple(cohort for cohort in SOURCE_COHORTS if cohort in observed)


def _path_is_relevant_candidate(path: Path, root: Path) -> bool:
    relative = _display_path(path, root).casefold()
    return any(token in relative for token in RELEVANCE_TOKENS)


def _forbidden_filename(name: str) -> bool:
    normalized = name.casefold()
    return normalized in FORBIDDEN_FILE_NAMES or any(
        token in normalized for token in FORBIDDEN_NAME_TOKENS
    )


def _validate_artifact_path(root: Path, path: Path) -> str | None:
    try:
        path.resolve().relative_to(root)
    except ValueError:
        return "artifact_outside_project_root"
    if path.is_symlink():
        return "artifact_symlink_forbidden"
    if not path.exists() or not path.is_file():
        return "artifact_missing"
    if path.suffix.casefold() not in ALLOWED_EXTENSIONS:
        return "artifact_extension_invalid"
    if _forbidden_filename(path.name):
        return "artifact_forbidden_sensitive_name"
    return None


def _artifact_type(extension: str) -> str:
    return {
        ".json": "json",
        ".jsonl": "jsonl",
        ".csv": "csv",
        ".parquet": "parquet",
        ".xlsx": "xlsx",
        ".sqlite": "sqlite",
        ".db": "sqlite",
        ".md": "documentation",
        ".txt": "text",
        ".yaml": "yaml",
        ".yml": "yaml",
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".webp": "image",
        ".pdf": "pdf",
        ".zip": "archive",
    }.get(extension, "unknown")


def _collect_mapping_keys(payload: Any, *, depth: int = 0) -> set[str]:
    if depth > 4:
        return set()
    keys: set[str] = set()
    if isinstance(payload, Mapping):
        for key, value in payload.items():
            keys.add(str(key))
            keys.update(_collect_mapping_keys(value, depth=depth + 1))
    elif isinstance(payload, list):
        for item in payload[:100]:
            keys.update(_collect_mapping_keys(item, depth=depth + 1))
    return keys


def _cohorts_from_payload(payload: Any, *, depth: int = 0) -> list[str]:
    if depth > 4:
        return []
    cohorts: set[str] = set()
    if isinstance(payload, Mapping):
        for key, value in payload.items():
            if str(key) in SOURCE_COHORT_COLUMNS:
                cohorts.update(_cohorts_from_text(str(value)))
            cohorts.update(_cohorts_from_payload(value, depth=depth + 1))
    elif isinstance(payload, list):
        for item in payload[:1000]:
            cohorts.update(_cohorts_from_payload(item, depth=depth + 1))
    return sorted(cohorts)


def _cohorts_from_text(value: str) -> list[str]:
    normalized = value.casefold()
    return [cohort for cohort in SOURCE_COHORTS if cohort in normalized]


def _declared_cohorts(value: Any) -> list[str]:
    values = value if isinstance(value, list) else [value]
    return sorted(
        {
            str(item)
            for item in values
            if item is not None and str(item) in SOURCE_COHORTS
        }
    )


def _first_text(mapping: Mapping[str, Any], fields: Sequence[str]) -> str | None:
    for field in fields:
        value = _optional_text(mapping.get(field))
        if value is not None:
            return value
    return None


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text.casefold() not in {"none", "null", "nan"} else None


def _optional_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _inventory_warnings(artifacts: Sequence[EvidenceArtifact]) -> list[str]:
    warnings: list[str] = []
    if any(artifact.sensitive_content_detected for artifact in artifacts):
        warnings.append("sensitive_artifacts_blocked")
    if any(artifact.artifact_type == "archive" for artifact in artifacts):
        warnings.append("archives_inventory_only_not_extracted")
    if any(artifact.artifact_type == "image" for artifact in artifacts):
        warnings.append("images_inventory_only_no_ocr_executed")
    return warnings


def _recommended_action(decision: str) -> str:
    return {
        "AUTHORITATIVE_EVIDENCE_COMPLETE_AND_JOINABLE": "DESIGN_VERSIONED_LEGACY_BRIDGE_CONTRACT",
        "PARTIAL_AUTHORITATIVE_EVIDENCE_FOUND": "TARGETED_EVIDENCE_RECOVERY_REQUIRED",
        "AUTHORITATIVE_EVIDENCE_NOT_JOINABLE": "SEGREGATE_LEGACY_MASTER_RESEARCH_ONLY",
        "CONFLICTING_AUTHORITATIVE_EVIDENCE": "RESOLVE_SOURCE_AUTHORITY_CONFLICTS",
        "NO_AUTHORITATIVE_EVIDENCE_FOUND": "SEGREGATE_LEGACY_MASTER_RESEARCH_ONLY",
    }[decision]


def _base_report(
    *,
    root: Path,
    trader_master_path: str | Path,
    source_profile_path: str | Path,
    evidence_roots: Sequence[str | Path] | None,
    write_report: bool,
    json_path: Path,
    markdown_path: Path,
    generated_at_utc: str | None,
) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": generated_at_utc or datetime.now(UTC).isoformat(),
        "status": "blocked",
        "reason": "not_evaluated",
        "decision": "NO_AUTHORITATIVE_EVIDENCE_FOUND",
        "fingerprint_spec_version": FINGERPRINT_SPEC_VERSION,
        "trader_master_path": _display_path(_resolve(root, trader_master_path), root),
        "source_profile_path": _display_path(_resolve(root, source_profile_path), root),
        "requested_evidence_roots": [
            _display_path(_resolve(root, item), root)
            for item in (evidence_roots or DEFAULT_EVIDENCE_ROOTS)
        ],
        "trader_master_sha256": None,
        "trader_master_row_count": 0,
        "source_cohorts": [],
        "evidence_roots_inspected": [],
        "discovered_file_count": 0,
        "ignored_forbidden_file_count": 0,
        "artifact_candidate_count": 0,
        "artifact_inspected_count": 0,
        "artifact_uninspected_count": 0,
        "artifact_blocked_count": 0,
        "inventory_accounting_consistent": True,
        "inventory_coverage_complete": True,
        "authoritative_evidence_absence_proven": False,
        "blocked_artifacts_may_contain_unassessed_evidence": False,
        "decision_scope": "complete_candidate_inventory",
        "evidence_inventory": [],
        "evidence_by_cohort_and_field": [],
        "authoritative_and_joinable_count": 0,
        "authoritative_but_not_joinable_count": 0,
        "informational_only_count": 0,
        "conflicting_count": 0,
        "missing_field_count": 0,
        "full_ocr_3141_resolvable_field_count": 0,
        "manual_queue_resolved_resolvable_field_count": 0,
        "all_required_fields_resolvable_by_cohort": {},
        "bridge_design_preconditions_satisfied": False,
        "recommended_next_action": "SEGREGATE_LEGACY_MASTER_RESEARCH_ONLY",
        "account_scope_hash_present": False,
        "account_scope_hash_valid": False,
        "account_scope_original_identifier_persisted": False,
        "import_eligible_true_count": 0,
        "write_requested": bool(write_report),
        "write_performed": False,
        "output_paths": {
            "json": _display_path(json_path, root),
            "markdown": _display_path(markdown_path, root),
        },
        "validation_errors": [],
        "blockers": [],
        "warnings": [],
        **SAFETY_FLAGS,
        "safety_flags": dict(SAFETY_FLAGS),
    }


def _blocked(
    report: dict[str, Any],
    reason: str,
    errors: Sequence[str] | None = None,
) -> dict[str, Any]:
    final = dict(report)
    blockers = sorted(set(errors or [reason]))
    final.update(
        status="blocked",
        reason=reason,
        validation_errors=blockers,
        blockers=blockers,
        write_performed=False,
        **SAFETY_FLAGS,
        safety_flags=dict(SAFETY_FLAGS),
    )
    return final


def _maybe_write(
    report: dict[str, Any],
    write_report: bool,
    json_path: Path,
    markdown_path: Path,
) -> dict[str, Any]:
    if not write_report:
        return report
    final = dict(report)
    final["write_performed"] = True
    _atomic_write(
        json_path,
        json.dumps(final, ensure_ascii=False, indent=2, sort_keys=True, default=str) + "\n",
    )
    _atomic_write(markdown_path, render_markdown(final))
    return final


def _atomic_write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{os.getpid()}.tmp")
    try:
        temporary.write_text(content, encoding="utf-8", newline="\n")
        temporary.replace(path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _validate_report_paths(root: Path, *paths: Path) -> list[str]:
    allowed_root = (root / "data/reports").resolve()
    errors: list[str] = []
    for path in paths:
        try:
            path.resolve().relative_to(allowed_root)
        except ValueError:
            errors.append(f"report_path_outside_data_reports:{path}")
        if path.suffix.casefold() not in {".json", ".md"}:
            errors.append(f"unsupported_report_extension:{path.suffix}")
    return sorted(set(errors))


def _resolve(root: Path, value: str | Path) -> Path:
    path = Path(value)
    return path.resolve() if path.is_absolute() else (root / path).resolve()


def _display_path(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return "<outside_project_root>"


def _display_path_without_resolve(path: Path, root: Path) -> str:
    try:
        return path.absolute().relative_to(root.absolute()).as_posix()
    except ValueError:
        return "<outside_project_root>"
