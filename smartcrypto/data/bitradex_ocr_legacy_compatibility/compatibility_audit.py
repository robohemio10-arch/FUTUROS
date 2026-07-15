"""Fail-closed read-only audit of the Bitradex OCR legacy contract."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import tempfile
from collections.abc import Callable, Mapping, Sequence
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from smartcrypto.data.trader_master_fingerprint_v2.master_adapter import (
    MasterReadBundle,
    read_trader_master_readonly,
)

from .contract import LegacyContract, LegacyContractError, load_legacy_contract


AUDIT_SCHEMA_VERSION = "bitradex_ocr_legacy_compatibility_audit_v1"
DECISION = "LEGACY_APPEND_CANDIDATE_REQUIRES_EXPLICIT_APPLY_AUTHORIZATION"
DEFAULT_CONTRACT_PATH = Path("config/bitradex_ocr_legacy_contract_v1.json")
DEFAULT_OUTPUT_JSON = Path("data/reports/bitradex_ocr_legacy_compatibility_v1.json")
DEFAULT_OUTPUT_MARKDOWN = Path("data/reports/bitradex_ocr_legacy_compatibility_v1.md")
EXPECTED_PIPELINE_STAGE = "PREVIEW_V4_RECONCILED_IMPORT_NOT_CONFIRMED"
XLSX_SHEET = "trades_master_candidate"

SAFETY_FLAGS: dict[str, bool] = {
    "operational_authority": False,
    "manual_authorization_required": True,
    "official_import_allowed": False,
    "import_executed": False,
    "writes_trader_master": False,
    "writes_xlsx": False,
    "writes_parquet": False,
    "writes_sqlite": False,
    "writes_runtime": False,
    "runs_qlib": False,
    "updates_ai_shadow": False,
    "changes_risk": False,
    "sends_orders": False,
    "exchange_private_access": False,
    "live_release_allowed": False,
    "canary_release_allowed": False,
}

MasterReader = Callable[..., MasterReadBundle]


# Each logical gate names concrete evidence locations. Values are never guessed from
# nearby fields. The second paths support the canonical nested Preview V4 shape.
PREVIEW_PATHS: dict[str, tuple[tuple[str, ...], ...]] = {
    "status": (("status",),),
    "package_token": (("package_token",),),
    "pipeline_stage": (("pipeline_stage",),),
    "incoming_rows": (("incoming_rows",), ("source", "incoming_row_count")),
    "master_reconciled_rows": (
        ("master_reconciled_rows",),
        ("sidecar_closeout", "reconciled_row_count"),
    ),
    "sidecar_auto_matched_rows": (
        ("sidecar_auto_matched_rows",),
        ("sidecar_closeout", "auto_matched_count"),
    ),
    "sidecar_residual_equivalence_rows": (
        ("sidecar_residual_equivalence_rows",),
        ("sidecar_closeout", "residual_equivalence_count"),
    ),
    "sidecars_reconciled": (
        ("sidecars_reconciled",),
        ("sidecar_closeout", "sidecars_reconciled"),
    ),
    "identity_contract_valid": (
        ("identity_contract_valid",),
        ("identity_contract", "identity_contract_valid"),
    ),
    "incoming_required_missing_total": (
        ("incoming_required_missing_total",),
        ("validation", "incoming_required_missing_total"),
    ),
    "incoming_internal_duplicate_excess_count": (
        ("incoming_internal_duplicate_excess_count",),
        ("validation", "incoming_strict_key_stats", "duplicate_excess_count"),
    ),
    "incoming_fallback_collision_count": (
        ("incoming_fallback_collision_count",),
        ("validation", "incoming_fallback_collision_count"),
    ),
    "exists_strict_count": (("exists_strict_count",), ("comparison", "exists_strict_count")),
    "exists_fallback_count": (
        ("exists_fallback_count",),
        ("comparison", "exists_fallback_count"),
    ),
    "exists_strict_duplicate_group_count": (
        ("exists_strict_duplicate_group_count",),
        ("comparison", "strict_duplicate_group_exists_count"),
    ),
    "novel_to_reconciled_master_count": (
        ("novel_to_reconciled_master_count",),
        ("comparison", "novel_to_reconciled_master_count"),
    ),
    "ambiguous_master_match_count": (
        ("ambiguous_master_match_count",),
        ("comparison", "ambiguous_master_match_count"),
    ),
    "invalid_identity_count": (
        ("invalid_identity_count",),
        ("comparison", "invalid_identity_count"),
    ),
    "source_image_conflict_count": (
        ("source_image_conflict_count",),
        ("validation", "source_image_conflict_count"),
    ),
    "guarded_apply_allowed": (
        ("guarded_apply_allowed",),
        ("gates", "guarded_apply_allowed"),
    ),
    "import_executed": (("import_executed",), ("official_import_executed",)),
    "master_preserved": (
        ("master_preserved",),
        ("gates", "master_source_hashes_preserved"),
    ),
    "stash_preserved": (("stash_preserved",), ("repository_after", "stash_preserved")),
    "worktree_clean": (("worktree_clean",), ("repository_after", "worktree_clean")),
}


def build_legacy_compatibility_audit(
    *,
    project_root: str | Path,
    contract_path: str | Path = DEFAULT_CONTRACT_PATH,
    preview_summary_path: str | Path | None = None,
    preview_csv_path: str | Path | None = None,
    master_xlsx_path: str | Path | None = None,
    master_parquet_path: str | Path | None = None,
    write_report: bool = False,
    output_json: str | Path = DEFAULT_OUTPUT_JSON,
    output_markdown: str | Path = DEFAULT_OUTPUT_MARKDOWN,
    master_reader: MasterReader = read_trader_master_readonly,
) -> dict[str, Any]:
    """Validate historical evidence without importing or changing source artifacts."""

    root = Path(project_root).resolve()
    json_path = _resolve_output(root, output_json)
    markdown_path = _resolve_output(root, output_markdown)
    report = _base_report(contract_path, write_report, json_path, markdown_path)

    output_errors = _validate_output_paths(root, json_path, markdown_path) if write_report else []
    if output_errors:
        return _finish(
            report,
            status="blocked",
            reason="unsafe_report_path",
            errors=output_errors,
            write_report=False,
            json_path=json_path,
            markdown_path=markdown_path,
        )

    contract_source, contract_path_error = _resolve_input(root, contract_path)
    if contract_path_error:
        return _finish(
            report,
            status="blocked",
            reason=contract_path_error,
            errors=[contract_path_error],
            write_report=write_report,
            json_path=json_path,
            markdown_path=markdown_path,
        )
    try:
        contract = load_legacy_contract(contract_source)
    except LegacyContractError as exc:
        return _finish(
            report,
            status="blocked",
            reason="legacy_contract_invalid",
            errors=[str(exc)],
            write_report=write_report,
            json_path=json_path,
            markdown_path=markdown_path,
        )

    _apply_contract(report, contract)
    resolved, path_errors = _resolve_contract_sources(
        root,
        contract,
        preview_summary_path=preview_summary_path,
        preview_csv_path=preview_csv_path,
        master_xlsx_path=master_xlsx_path,
        master_parquet_path=master_parquet_path,
    )
    report["resolved_sources"] = {key: str(value) for key, value in resolved.items()}
    if path_errors:
        return _finish(
            report,
            status="blocked",
            reason="source_path_validation_failed",
            errors=path_errors,
            write_report=write_report,
            json_path=json_path,
            markdown_path=markdown_path,
        )

    errors: list[str] = []
    preview_payload = _load_json_object(resolved["preview_summary"], errors)
    preview_values: dict[str, Any] = {}
    preview_evidence: dict[str, Any] = {}
    if preview_payload is not None:
        preview_values, preview_evidence, preview_errors = _validate_preview(preview_payload, contract)
        errors.extend(preview_errors)
    report["preview_values"] = preview_values
    report["preview_field_evidence"] = preview_evidence
    report["preview_v4_verified"] = not any(error.startswith("preview_") for error in errors)

    csv_rows = _count_csv_rows(resolved["preview_csv"], errors)
    report["preview_csv_row_count"] = csv_rows
    if csv_rows != contract.expected_counts.retained_candidate_rows:
        errors.append(
            "preview_csv_row_count_mismatch:"
            f"expected={contract.expected_counts.retained_candidate_rows}:actual={csv_rows}"
        )

    parquet_result = _audit_parquet_master(
        root, resolved["master_parquet"], contract, master_reader=master_reader
    )
    errors.extend(parquet_result.pop("_validation_errors", []))
    report.update(parquet_result)

    xlsx_result = _audit_xlsx_master(resolved["master_xlsx"], contract)
    errors.extend(xlsx_result.pop("_validation_errors", []))
    report.update(xlsx_result)

    report["legacy_contract_compatible"] = not errors
    reason = (
        "legacy_contract_compatibility_confirmed"
        if not errors
        else "legacy_contract_compatibility_blocked"
    )
    return _finish(
        report,
        status="ok" if not errors else "blocked",
        reason=reason,
        errors=errors,
        write_report=write_report,
        json_path=json_path,
        markdown_path=markdown_path,
    )


def _apply_contract(report: dict[str, Any], contract: LegacyContract) -> None:
    counts = contract.expected_counts
    funding = contract.funding_policy
    identity = contract.identity_policy
    report.update(
        contract_id=contract.contract_id,
        batch_id=contract.batch_id,
        contract_mode=contract.contract_mode,
        legacy_append_candidate_count=counts.retained_candidate_rows,
        master_rows_before=counts.master_rows_before,
        expected_master_rows_after_authorized_append=counts.master_rows_after_authorized_append,
        funding_required_for_legacy_contract=funding.funding_required_for_legacy_contract,
        funding_status=funding.funding_status,
        funding_fee_value=funding.funding_fee_value,
        funding_assumed_zero=funding.funding_assumed_zero,
        funding_derived_as_residual=funding.funding_derived_as_residual,
        v2_financial_decomposition_eligible=funding.v2_financial_decomposition_eligible,
        synthetic_order_id_authoritative=identity.synthetic_order_id_authoritative,
        synthetic_order_id_role=identity.synthetic_order_id_role,
        v2_primary_identity_eligible=identity.v2_primary_identity_eligible,
    )


def _resolve_contract_sources(
    root: Path,
    contract: LegacyContract,
    **overrides: str | Path | None,
) -> tuple[dict[str, Path], list[str]]:
    defaults = {
        "preview_summary": contract.sources.preview_summary,
        "preview_csv": contract.sources.preview_csv,
        "master_xlsx": contract.sources.master_xlsx,
        "master_parquet": contract.sources.master_parquet,
    }
    resolved: dict[str, Path] = {}
    errors: list[str] = []
    for name, default in defaults.items():
        requested = overrides.get(f"{name}_path") or default
        path, error = _resolve_input(root, requested)
        resolved[name] = path
        if error:
            errors.append(f"{name}:{error}")
    return resolved, sorted(errors)


def _validate_preview(
    payload: Mapping[str, Any], contract: LegacyContract
) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
    counts = contract.expected_counts
    expected: dict[str, Any] = {
        "status": "ok",
        "package_token": contract.batch_id,
        "pipeline_stage": EXPECTED_PIPELINE_STAGE,
        "incoming_rows": counts.retained_candidate_rows,
        "master_reconciled_rows": counts.master_rows_before,
        "sidecar_auto_matched_rows": counts.sidecar_auto_matched_rows,
        "sidecar_residual_equivalence_rows": counts.sidecar_residual_equivalence_rows,
        "sidecars_reconciled": True,
        "identity_contract_valid": True,
        "incoming_required_missing_total": 0,
        "incoming_internal_duplicate_excess_count": counts.internal_duplicate_excess_rows,
        "incoming_fallback_collision_count": counts.fallback_collision_rows,
        "exists_strict_count": 0,
        "exists_fallback_count": 0,
        "exists_strict_duplicate_group_count": 0,
        "novel_to_reconciled_master_count": counts.legacy_novel_candidate_rows,
        "ambiguous_master_match_count": counts.ambiguous_master_match_rows,
        "invalid_identity_count": counts.invalid_identity_rows,
        "source_image_conflict_count": counts.source_image_conflict_rows,
        "guarded_apply_allowed": False,
        "import_executed": False,
        "master_preserved": True,
        "stash_preserved": True,
        "worktree_clean": True,
    }
    values: dict[str, Any] = {}
    evidence: dict[str, Any] = {}
    errors: list[str] = []
    for field, expected_value in expected.items():
        found, value, source_path = _extract_preview_value(payload, PREVIEW_PATHS[field])
        if field == "pipeline_stage" and not found:
            found, value, source_path = _normalize_legacy_pipeline_stage(payload)
        if not found:
            errors.append(f"preview_required_field_missing:{field}")
            evidence[field] = {"found": False, "source_path": None}
            continue
        values[field] = value
        evidence[field] = {"found": True, "source_path": source_path}
        if type(value) is not type(expected_value) or value != expected_value:
            errors.append(
                f"preview_value_mismatch:{field}:expected={expected_value!r}:actual={value!r}"
            )
    return values, evidence, sorted(errors)


def _normalize_legacy_pipeline_stage(
    payload: Mapping[str, Any],
) -> tuple[bool, Any, str | None]:
    """Normalize the canonical V4 completion marker into its explicit stage enum."""

    found_reason, reason, reason_path = _extract_preview_value(payload, (("reason",),))
    found_preview, preview_only, preview_path = _extract_preview_value(payload, (("preview_only",),))
    found_import, import_executed, import_path = _extract_preview_value(
        payload, (("official_import_executed",),)
    )
    if (
        found_reason
        and reason == "bitradex_ocr_import_preview_v4_completed"
        and found_preview
        and preview_only is True
        and found_import
        and import_executed is False
    ):
        evidence = "+".join(path for path in (reason_path, preview_path, import_path) if path)
        return True, EXPECTED_PIPELINE_STAGE, f"legacy_v4_stage_marker:{evidence}"
    return False, None, None


def _extract_preview_value(
    payload: Mapping[str, Any], paths: Sequence[Sequence[str]]
) -> tuple[bool, Any, str | None]:
    for path in paths:
        current: Any = payload
        used: list[str] = []
        for requested_key in path:
            if not isinstance(current, Mapping):
                break
            matches = [key for key in current if _normalize_key(str(key)) == _normalize_key(requested_key)]
            if len(matches) != 1:
                break
            actual_key = matches[0]
            used.append(str(actual_key))
            current = current[actual_key]
        else:
            return True, current, ".".join(used)
    return False, None, None


def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")


def _audit_parquet_master(
    root: Path,
    path: Path,
    contract: LegacyContract,
    *,
    master_reader: MasterReader,
) -> dict[str, Any]:
    errors: list[str] = []
    bundle = master_reader(project_root=root, trader_master_path=path)
    source_report = bundle.report
    actual_hash = source_report.get("trader_master_sha256_before")
    expected_hash = contract.expected_master_hashes.parquet_sha256
    row_count = source_report.get("trader_master_row_count", 0)
    columns = source_report.get("trader_master_schema_columns", [])
    hash_preserved = bool(source_report.get("trader_master_hash_preserved"))
    if source_report.get("status") != "ok":
        errors.append(f"master_parquet_read_blocked:{source_report.get('reason')}")
    if actual_hash != expected_hash:
        errors.append(f"master_parquet_sha256_mismatch:expected={expected_hash}:actual={actual_hash}")
    if row_count != contract.expected_counts.master_rows_before:
        errors.append(
            "master_parquet_row_count_mismatch:"
            f"expected={contract.expected_counts.master_rows_before}:actual={row_count}"
        )
    schema_verified = list(columns) == list(contract.historical_master_schema.columns)
    if not schema_verified:
        errors.append("master_parquet_schema_mismatch")
    if not hash_preserved:
        errors.append("master_parquet_hash_not_preserved")
    return {
        "master_parquet_hash_preserved": hash_preserved,
        "master_parquet_row_count": row_count,
        "master_parquet_sha256": actual_hash,
        "master_parquet_temp_copy_used": bool(
            source_report.get("trader_master_temp_copy_used")
        ),
        "master_schema_verified": schema_verified,
        "master_schema_columns": list(columns),
        "_validation_errors": errors,
    }


def _audit_xlsx_master(path: Path, contract: LegacyContract) -> dict[str, Any]:
    errors: list[str] = []
    expected_hash = contract.expected_master_hashes.xlsx_sha256
    try:
        size_before = path.stat().st_size
        sha_before = _file_sha256(path)
        with TemporaryDirectory(prefix="legacy-master-xlsx-readonly-") as temporary:
            copy = Path(temporary) / "trades_master.xlsx"
            shutil.copy2(path, copy)
            row_count = _read_xlsx_row_count(copy)
        size_after = path.stat().st_size
        sha_after = _file_sha256(path)
    except Exception as exc:
        return {
            "master_xlsx_hash_preserved": False,
            "master_xlsx_row_count": 0,
            "master_xlsx_sha256": None,
            "master_xlsx_temp_copy_used": True,
            "_validation_errors": [f"master_xlsx_unreadable:{type(exc).__name__}"],
        }
    hash_preserved = sha_before == sha_after and size_before == size_after
    if sha_before != expected_hash:
        errors.append(f"master_xlsx_sha256_mismatch:expected={expected_hash}:actual={sha_before}")
    if row_count != contract.expected_counts.master_rows_before:
        errors.append(
            "master_xlsx_row_count_mismatch:"
            f"expected={contract.expected_counts.master_rows_before}:actual={row_count}"
        )
    if not hash_preserved:
        errors.append("master_xlsx_hash_not_preserved")
    return {
        "master_xlsx_hash_preserved": hash_preserved,
        "master_xlsx_row_count": row_count,
        "master_xlsx_sha256": sha_before,
        "master_xlsx_temp_copy_used": True,
        "_validation_errors": errors,
    }


def _read_xlsx_row_count(path: Path) -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if XLSX_SHEET not in workbook.sheetnames:
            raise ValueError("master_xlsx_sheet_missing")
        rows = workbook[XLSX_SHEET].iter_rows(values_only=True)
        next(rows, None)
        return sum(1 for values in rows if any(value is not None for value in values))
    finally:
        workbook.close()


def _count_csv_rows(path: Path, errors: list[str]) -> int:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, None)
            if not header:
                errors.append("preview_csv_header_missing")
                return 0
            return sum(1 for row in reader if any(cell.strip() for cell in row))
    except (OSError, UnicodeError, csv.Error) as exc:
        errors.append(f"preview_csv_unreadable:{type(exc).__name__}")
        return 0


def _load_json_object(path: Path, errors: list[str]) -> Mapping[str, Any] | None:
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        errors.append(f"preview_summary_unreadable:{type(exc).__name__}")
        return None
    if not isinstance(payload, dict):
        errors.append("preview_summary_root_must_be_object")
        return None
    return payload


def _base_report(
    contract_path: str | Path, write_report: bool, json_path: Path, markdown_path: Path
) -> dict[str, Any]:
    return {
        "schema_version": AUDIT_SCHEMA_VERSION,
        "status": "blocked",
        "reason": "not_evaluated",
        "decision": DECISION,
        "contract_path": str(contract_path),
        "contract_id": None,
        "batch_id": None,
        "legacy_contract_compatible": False,
        "legacy_append_candidate_count": 0,
        "master_rows_before": 0,
        "expected_master_rows_after_authorized_append": 0,
        "preview_v4_verified": False,
        "preview_csv_row_count": 0,
        "master_xlsx_hash_preserved": False,
        "master_parquet_hash_preserved": False,
        "master_xlsx_row_count": 0,
        "master_parquet_row_count": 0,
        "master_schema_verified": False,
        "validation_errors": [],
        "write_requested": bool(write_report),
        "write_performed": False,
        "output_json": str(json_path),
        "output_markdown": str(markdown_path),
        **SAFETY_FLAGS,
        "safety_flags": dict(SAFETY_FLAGS),
    }


def _finish(
    report: dict[str, Any],
    *,
    status: str,
    reason: str,
    errors: Sequence[str],
    write_report: bool,
    json_path: Path,
    markdown_path: Path,
) -> dict[str, Any]:
    final = dict(report)
    final.update(status=status, reason=reason, validation_errors=sorted(set(errors)))
    if write_report:
        final["write_performed"] = True
        _atomic_write_text(
            json_path,
            json.dumps(final, ensure_ascii=False, indent=2, sort_keys=True, default=str) + "\n",
        )
        _atomic_write_text(markdown_path, render_markdown(final))
    return final


def render_markdown(report: Mapping[str, Any]) -> str:
    lines = [
        "# Bitradex OCR Legacy Contract Compatibility V1",
        "",
        f"- Status: `{report.get('status')}`",
        f"- Reason: `{report.get('reason')}`",
        f"- Decision: `{report.get('decision')}`",
        f"- Contract: `{report.get('contract_id')}`",
        f"- Batch: `{report.get('batch_id')}`",
        f"- Legacy candidates: `{report.get('legacy_append_candidate_count', 0)}`",
        f"- Master rows before: `{report.get('master_rows_before', 0)}`",
        "- Funding: unknown and unavailable in the source; it is not zero or a residual.",
        "- Synthetic order IDs: lineage aliases only, never native financial identity.",
        "",
        "## Authority boundary",
        "",
        "This is read-only compatibility evidence. It does not authorize or execute an import.",
        "A separate apply branch requires explicit authorization, backup, final preview, atomic",
        "write, and post-import audit.",
        "",
        "## Validation errors",
        "",
    ]
    errors = report.get("validation_errors") or []
    lines.extend(f"- `{error}`" for error in errors)
    if not errors:
        lines.append("- None")
    lines.append("")
    return "\n".join(lines)


def _resolve_input(root: Path, value: str | Path) -> tuple[Path, str | None]:
    requested = Path(value)
    candidate = requested if requested.is_absolute() else root / requested
    if _has_symlink_component(candidate):
        return candidate, "input_symlink_rejected"
    resolved = candidate.resolve()
    if not requested.is_absolute():
        try:
            resolved.relative_to(root)
        except ValueError:
            return resolved, "relative_input_outside_project_root"
    if not resolved.is_file():
        return resolved, "input_file_missing"
    return resolved, None


def _resolve_output(root: Path, value: str | Path) -> Path:
    requested = Path(value)
    return requested.resolve() if requested.is_absolute() else (root / requested).resolve()


def _validate_output_paths(root: Path, json_path: Path, markdown_path: Path) -> list[str]:
    allowed_root = (root / "data" / "reports").resolve()
    errors: list[str] = []
    for path, extension in ((json_path, ".json"), (markdown_path, ".md")):
        try:
            path.relative_to(allowed_root)
        except ValueError:
            errors.append(f"report_path_outside_data_reports:{path}")
        if path.suffix.casefold() != extension:
            errors.append(f"report_extension_invalid:{path}")
        if _has_symlink_component(path):
            errors.append(f"report_symlink_rejected:{path}")
    return sorted(errors)


def _has_symlink_component(path: Path) -> bool:
    current = path
    while True:
        if current.is_symlink():
            return True
        if current.parent == current:
            return False
        current = current.parent


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        temporary.replace(path)
    finally:
        if temporary.exists():
            temporary.unlink()
