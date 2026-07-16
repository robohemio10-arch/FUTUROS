"""Filesystem primitives for the canonical XLSX two-master transaction."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook, load_workbook

from .contract import TransitionContract, file_sha256


@dataclass(frozen=True)
class BackupEvidence:
    directory: Path
    xlsx_path: Path
    parquet_path: Path
    xlsx_sha256: str
    parquet_sha256: str
    xlsx_size: int
    parquet_size: int


@dataclass(frozen=True)
class TargetEvidence:
    xlsx_path: Path
    parquet_path: Path
    xlsx_row_count: int
    parquet_row_count: int
    xlsx_semantic_sha256: str
    parquet_semantic_sha256: str
    prefix_semantic_sha256: str
    tail_semantic_sha256: str
    xlsx_sha256: str
    parquet_sha256: str
    xlsx_size: int
    parquet_size: int


class TransitionLock:
    def __init__(self, path: Path, transition_id: str) -> None:
        self.path = path
        self.transition_id = transition_id
        self.created = False

    def acquire(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        descriptor = os.open(
            self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600
        )
        payload = {
            "transition_id": self.transition_id,
            "pid": os.getpid(),
            "created_at_utc": datetime.now(UTC).isoformat(),
        }
        try:
            os.write(
                descriptor,
                (json.dumps(payload, sort_keys=True) + "\n").encode("utf-8"),
            )
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
        self.created = True

    def release(self) -> None:
        if self.created:
            self.path.unlink(missing_ok=True)
            self.created = False


def create_verified_backups(
    *,
    root: Path,
    contract: TransitionContract,
    run_id: str,
) -> BackupEvidence:
    directory = (
        root
        / "data/backups/bitradex_ocr_legacy_canonical_xlsx_transition_v2"
        / run_id
    )
    directory.mkdir(parents=True, exist_ok=False)
    source_xlsx = root / contract.pre_state.master_xlsx_path
    source_parquet = root / contract.pre_state.master_parquet_path
    xlsx = directory / source_xlsx.name
    parquet = directory / source_parquet.name
    _copy_fsync(source_xlsx, xlsx)
    _copy_fsync(source_parquet, parquet)
    evidence = BackupEvidence(
        directory=directory,
        xlsx_path=xlsx,
        parquet_path=parquet,
        xlsx_sha256=file_sha256(xlsx),
        parquet_sha256=file_sha256(parquet),
        xlsx_size=xlsx.stat().st_size,
        parquet_size=parquet.stat().st_size,
    )
    if evidence.xlsx_sha256 != contract.pre_state.master_xlsx_sha256:
        raise RuntimeError("xlsx_backup_hash_mismatch")
    if evidence.parquet_sha256 != contract.pre_state.master_parquet_sha256:
        raise RuntimeError("parquet_backup_hash_mismatch")
    if (
        evidence.xlsx_size != source_xlsx.stat().st_size
        or evidence.parquet_size != source_parquet.stat().st_size
    ):
        raise RuntimeError("backup_size_mismatch")
    return evidence


def build_verified_targets(
    *,
    master_parquet: Path,
    destination_directory: Path,
    contract: TransitionContract,
    candidates: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
    semantic_hasher: Any,
) -> TargetEvidence:
    """Build both target formats from the canonical Parquet prefix."""

    destination_directory.mkdir(parents=True, exist_ok=True)
    parquet_target = destination_directory / "trades_master.canonical.parquet"
    xlsx_target = destination_directory / "trades_master.canonical.xlsx"
    master = pq.read_table(master_parquet)
    if tuple(master.column_names) != tuple(columns):
        raise RuntimeError("parquet_master_schema_mismatch")
    source_rows = [dict(row) for row in master.to_pylist()]
    if len(source_rows) != contract.pre_state.master_row_count:
        raise RuntimeError("parquet_master_row_count_mismatch")

    target_schema = pa.schema(
        [pa.field(column, pa.string()) for column in columns]
    )
    candidate_table = pa.Table.from_pylist(
        [
            {
                column: (
                    None
                    if row.get(column) is None
                    else str(row.get(column))
                )
                for column in columns
            }
            for row in candidates
        ],
        schema=target_schema,
    )
    for index, column in enumerate(columns):
        source_nonempty = sum(
            row.get(column) not in (None, "") for row in candidates
        )
        result_nonnull = (
            candidate_table.column(index).length()
            - candidate_table.column(index).null_count
        )
        if result_nonnull < source_nonempty:
            raise RuntimeError(f"silent_null_coercion:{column}")
    master_table = pa.Table.from_pylist(
        [
            {
                column: (
                    None
                    if row.get(column) is None
                    else str(row.get(column))
                )
                for column in columns
            }
            for row in source_rows
        ],
        schema=target_schema,
    )
    combined = pa.concat_tables([master_table, candidate_table])
    pq.write_table(combined, parquet_target)
    _fsync_file(parquet_target)
    parquet_rows = _read_parquet_rows(parquet_target, columns)

    _write_canonical_xlsx(
        xlsx_target,
        parquet_rows,
        columns,
        contract.target_state.canonical_xlsx_sheet,
    )
    xlsx_rows = _read_xlsx_rows(
        xlsx_target,
        columns,
        contract.target_state.canonical_xlsx_sheet,
    )
    expected = contract.target_state.expected_row_count
    if len(parquet_rows) != expected or len(xlsx_rows) != expected:
        raise RuntimeError("target_row_count_mismatch")
    prefix_length = contract.pre_state.master_row_count
    prefix_hash = semantic_hasher(parquet_rows[:prefix_length], columns)
    tail_hash = semantic_hasher(parquet_rows[prefix_length:], columns)
    target_parquet_hash = semantic_hasher(parquet_rows, columns)
    target_xlsx_hash = semantic_hasher(xlsx_rows, columns)
    if prefix_hash != semantic_hasher(source_rows, columns):
        raise RuntimeError("target_prefix_semantic_mismatch")
    if tail_hash != semantic_hasher(candidates, columns):
        raise RuntimeError("target_tail_semantic_mismatch")
    if target_parquet_hash != target_xlsx_hash:
        raise RuntimeError("target_cross_format_semantic_mismatch")
    if prefix_hash != contract.target_state.expected_prefix_semantic_sha256:
        raise RuntimeError("target_prefix_contract_mismatch")
    if tail_hash != contract.target_state.expected_tail_semantic_sha256:
        raise RuntimeError("target_tail_contract_mismatch")
    if (
        target_parquet_hash
        != contract.target_state.expected_target_semantic_sha256
    ):
        raise RuntimeError("target_semantic_contract_mismatch")
    return TargetEvidence(
        xlsx_path=xlsx_target,
        parquet_path=parquet_target,
        xlsx_row_count=len(xlsx_rows),
        parquet_row_count=len(parquet_rows),
        xlsx_semantic_sha256=target_xlsx_hash,
        parquet_semantic_sha256=target_parquet_hash,
        prefix_semantic_sha256=prefix_hash,
        tail_semantic_sha256=tail_hash,
        xlsx_sha256=file_sha256(xlsx_target),
        parquet_sha256=file_sha256(parquet_target),
        xlsx_size=xlsx_target.stat().st_size,
        parquet_size=parquet_target.stat().st_size,
    )


def restore_from_backups(
    *,
    master_xlsx: Path,
    master_parquet: Path,
    backup: BackupEvidence,
) -> bool:
    xlsx_restore = _temporary_peer(master_xlsx, ".rollback.xlsx")
    parquet_restore = _temporary_peer(
        master_parquet, ".rollback.parquet"
    )
    try:
        _copy_fsync(backup.xlsx_path, xlsx_restore)
        _copy_fsync(backup.parquet_path, parquet_restore)
        os.replace(parquet_restore, master_parquet)
        os.replace(xlsx_restore, master_xlsx)
        return (
            file_sha256(master_xlsx) == backup.xlsx_sha256
            and file_sha256(master_parquet) == backup.parquet_sha256
        )
    finally:
        xlsx_restore.unlink(missing_ok=True)
        parquet_restore.unlink(missing_ok=True)


def inspect_legacy_workbook(
    path: Path,
    *,
    data_sheet: str,
    summary_sheet: str,
    expected_header: Sequence[str],
) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
        if data_sheet not in sheet_names:
            raise RuntimeError("legacy_xlsx_data_sheet_missing")
        if summary_sheet not in sheet_names:
            raise RuntimeError("legacy_xlsx_build_summary_missing")
        iterator = workbook[data_sheet].iter_rows(values_only=True)
        header = tuple(str(value) if value is not None else "" for value in next(iterator, ()))
        if header != tuple(expected_header):
            raise RuntimeError("legacy_xlsx_header_mismatch")
        row_count = sum(
            1 for row in iterator if any(value is not None for value in row)
        )
        return {
            "sheet_names": sheet_names,
            "data_sheet": data_sheet,
            "summary_sheet": summary_sheet,
            "header": list(header),
            "header_column_count": len(header),
            "data_row_count": row_count,
            "layout_verified": True,
            "classification": "legacy_ocr_evidence_workbook",
        }
    finally:
        workbook.close()


def _write_canonical_xlsx(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
    sheet_name: str,
) -> None:
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "SMART FUTUROS"
    workbook.properties.title = "Canonical legacy Trader Master"
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(list(columns))
    for row in rows:
        worksheet.append([row.get(column) for column in columns])
    workbook.save(path)
    _fsync_file(path)


def _read_parquet_rows(
    path: Path, columns: Sequence[str]
) -> list[dict[str, Any]]:
    table = pq.read_table(path, columns=list(columns))
    return [dict(row) for row in table.to_pylist()]


def _read_xlsx_rows(
    path: Path, columns: Sequence[str], sheet_name: str
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [sheet_name]:
            raise RuntimeError("canonical_xlsx_sheet_set_invalid")
        iterator = workbook[sheet_name].iter_rows(values_only=True)
        header = tuple(next(iterator, ()))
        if header != tuple(columns):
            raise RuntimeError("canonical_xlsx_header_mismatch")
        return [
            dict(zip(columns, values, strict=True))
            for values in iterator
            if any(value is not None for value in values)
        ]
    finally:
        workbook.close()


def _copy_fsync(source: Path, destination: Path) -> None:
    shutil.copy2(source, destination)
    _fsync_file(destination)


def _fsync_file(path: Path) -> None:
    with path.open("r+b") as handle:
        handle.flush()
        os.fsync(handle.fileno())


def _temporary_peer(destination: Path, suffix: str) -> Path:
    descriptor, name = tempfile.mkstemp(
        prefix=f".{destination.name}.",
        suffix=suffix,
        dir=destination.parent,
    )
    os.close(descriptor)
    path = Path(name)
    path.unlink()
    return path
