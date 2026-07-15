from __future__ import annotations

import csv
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from smartcrypto.data.bitradex_ocr_legacy_compatibility.compatibility_audit import (
    AUDIT_SCHEMA_VERSION,
    DECISION,
    SAFETY_FLAGS,
    build_legacy_compatibility_audit,
)
from smartcrypto.data.bitradex_ocr_legacy_compatibility.contract import (
    LegacyContractError,
    load_legacy_contract,
)


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "audit_bitradex_ocr_legacy_compatibility_v1.py"
CONTRACT_SOURCE = ROOT / "config" / "bitradex_ocr_legacy_contract_v1.json"
MASTER_COLUMNS = json.loads(CONTRACT_SOURCE.read_text(encoding="utf-8"))[
    "historical_master_schema"
]["columns"]
ADAPTER_V2_SHA256 = "118bdf3d85814ac1b6d2f2a670e98ae9a724a63d97ce889b32de7494418e9323"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def valid_preview() -> dict[str, Any]:
    return {
        "status": "ok",
        "package_token": "20260714_151816",
        "pipeline_stage": "PREVIEW_V4_RECONCILED_IMPORT_NOT_CONFIRMED",
        "incoming_rows": 504,
        "master_reconciled_rows": 3058,
        "sidecar_auto_matched_rows": 3057,
        "sidecar_residual_equivalence_rows": 1,
        "sidecars_reconciled": True,
        "identity_contract_valid": True,
        "incoming_required_missing_total": 0,
        "incoming_internal_duplicate_excess_count": 0,
        "incoming_fallback_collision_count": 0,
        "exists_strict_count": 0,
        "exists_fallback_count": 0,
        "exists_strict_duplicate_group_count": 0,
        "novel_to_reconciled_master_count": 504,
        "ambiguous_master_match_count": 0,
        "invalid_identity_count": 0,
        "source_image_conflict_count": 0,
        "guarded_apply_allowed": False,
        "import_executed": False,
        "master_preserved": True,
        "stash_preserved": True,
        "worktree_clean": True,
    }


def nested_v4_preview() -> dict[str, Any]:
    return {
        "status": "ok",
        "reason": "bitradex_ocr_import_preview_v4_completed",
        "package_token": "20260714_151816",
        "preview_only": True,
        "official_import_executed": False,
        "source": {"incoming_row_count": 504},
        "sidecar_closeout": {
            "reconciled_row_count": 3058,
            "auto_matched_count": 3057,
            "residual_equivalence_count": 1,
            "sidecars_reconciled": True,
        },
        "identity_contract": {"identity_contract_valid": True},
        "validation": {
            "incoming_required_missing_total": 0,
            "incoming_strict_key_stats": {"duplicate_excess_count": 0},
            "incoming_fallback_collision_count": 0,
            "source_image_conflict_count": 0,
        },
        "comparison": {
            "exists_strict_count": 0,
            "exists_fallback_count": 0,
            "strict_duplicate_group_exists_count": 0,
            "novel_to_reconciled_master_count": 504,
            "ambiguous_master_match_count": 0,
            "invalid_identity_count": 0,
        },
        "gates": {
            "guarded_apply_allowed": False,
            "master_source_hashes_preserved": True,
        },
        "repository_after": {"stash_preserved": True, "worktree_clean": True},
    }


@pytest.fixture(scope="module")
def project_template(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("legacy-contract-template")
    package = root / "data/staging/bitradex_ocr/package_20260714_151816"
    trades = root / "data/trades"
    config = root / "config"
    package.mkdir(parents=True)
    trades.mkdir(parents=True)
    config.mkdir(parents=True)

    preview_summary = package / "BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json"
    preview_summary.write_text(json.dumps(valid_preview()), encoding="utf-8")
    preview_csv = package / "BITRADEX_OCR_IMPORT_PREVIEW_V4.csv"
    with preview_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["synthetic_order_id", "source_file"])
        writer.writerows((f"alias-{index}", f"image-{index}.jpg") for index in range(504))

    frame = pd.DataFrame({column: [None] * 3058 for column in MASTER_COLUMNS})
    frame["moeda"] = ["BTCUSDT"] * 3058
    frame.to_parquet(trades / "trades_master.parquet", index=False)

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("trades_master_candidate")
    worksheet.append(MASTER_COLUMNS)
    empty_tail = [None] * (len(MASTER_COLUMNS) - 1)
    for _ in range(3058):
        worksheet.append(["BTCUSDT", *empty_tail])
    workbook.save(trades / "trades_master.xlsx")

    contract = json.loads(CONTRACT_SOURCE.read_text(encoding="utf-8"))
    contract["expected_master_hashes"] = {
        "xlsx_sha256": sha256(trades / "trades_master.xlsx"),
        "parquet_sha256": sha256(trades / "trades_master.parquet"),
    }
    (config / "bitradex_ocr_legacy_contract_v1.json").write_text(
        json.dumps(contract), encoding="utf-8"
    )
    return root


def copy_project(tmp_path: Path, project_template: Path) -> Path:
    target = tmp_path / "project"
    shutil.copytree(project_template, target)
    return target


def contract_path(root: Path) -> Path:
    return root / "config/bitradex_ocr_legacy_contract_v1.json"


def mutate_contract(root: Path, *keys: str, value: Any) -> None:
    path = contract_path(root)
    payload = json.loads(path.read_text(encoding="utf-8"))
    current = payload
    for key in keys[:-1]:
        current = current[key]
    current[keys[-1]] = value
    path.write_text(json.dumps(payload), encoding="utf-8")


def audit(root: Path, **kwargs: Any) -> dict[str, Any]:
    return build_legacy_compatibility_audit(project_root=root, **kwargs)


def test_valid_contract_loads(project_template: Path) -> None:
    contract = load_legacy_contract(contract_path(project_template))
    assert contract.expected_counts.retained_candidate_rows == 504
    assert contract.funding_policy.funding_fee_value is None


@pytest.mark.parametrize(
    ("keys", "value", "error"),
    [
        (("schema_version",), "wrong", "schema_version_invalid"),
        (("expected_master_hashes", "xlsx_sha256"), "bad", "contract_sha256_invalid"),
        (("funding_policy", "funding_fee_value"), 0, "funding_fee_value_must_be_null"),
        (("funding_policy", "funding_assumed_zero"), True, "funding_assumed_zero_must_be_false"),
        (
            ("funding_policy", "funding_derived_as_residual"),
            True,
            "funding_derived_as_residual_must_be_false",
        ),
        (
            ("identity_policy", "synthetic_order_id_authoritative"),
            True,
            "synthetic_order_id_authoritative_must_be_false",
        ),
        (
            ("identity_policy", "account_scope_required_for_legacy_contract"),
            True,
            "account_scope_required_for_legacy_contract_must_be_false",
        ),
        (("authority", "import_authorized"), True, "import_authorized_must_be_false"),
        (("authority", "write_authorized"), True, "write_authorized_must_be_false"),
    ],
)
def test_unsafe_contract_values_are_rejected(
    tmp_path: Path,
    project_template: Path,
    keys: tuple[str, ...],
    value: Any,
    error: str,
) -> None:
    root = copy_project(tmp_path, project_template)
    mutate_contract(root, *keys, value=value)
    with pytest.raises(LegacyContractError, match=error):
        load_legacy_contract(contract_path(root))


def test_506_2_504_count_mismatch_is_rejected(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    mutate_contract(root, "expected_counts", "excluded_exact_duplicates", value=3)
    with pytest.raises(LegacyContractError, match="ocr_duplicate_retained_count_inconsistent"):
        load_legacy_contract(contract_path(root))


def test_3058_504_3562_count_mismatch_is_rejected(
    tmp_path: Path, project_template: Path
) -> None:
    root = copy_project(tmp_path, project_template)
    mutate_contract(root, "expected_counts", "master_rows_after_authorized_append", value=3563)
    with pytest.raises(LegacyContractError, match="master_authorized_append_count_inconsistent"):
        load_legacy_contract(contract_path(root))


def test_valid_preview_v4_returns_ok(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    report = audit(root)
    assert report["status"] == "ok"
    assert report["reason"] == "legacy_contract_compatibility_confirmed"
    assert report["decision"] == DECISION
    assert report["preview_v4_verified"] is True


def test_nested_real_v4_shape_is_verified(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    summary = root / "data/staging/bitradex_ocr/package_20260714_151816/BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json"
    summary.write_text(json.dumps(nested_v4_preview()), encoding="utf-8")
    report = audit(root)
    assert report["status"] == "ok"
    assert report["preview_values"]["pipeline_stage"] == (
        "PREVIEW_V4_RECONCILED_IMPORT_NOT_CONFIRMED"
    )
    assert report["preview_field_evidence"]["incoming_rows"]["source_path"] == (
        "source.incoming_row_count"
    )


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("novel_to_reconciled_master_count", 503),
        ("ambiguous_master_match_count", 1),
        ("sidecars_reconciled", False),
    ],
)
def test_preview_gate_mismatch_blocks(
    tmp_path: Path, project_template: Path, field: str, value: Any
) -> None:
    root = copy_project(tmp_path, project_template)
    summary = root / "data/staging/bitradex_ocr/package_20260714_151816/BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json"
    payload = valid_preview()
    payload[field] = value
    summary.write_text(json.dumps(payload), encoding="utf-8")
    report = audit(root)
    assert report["status"] == "blocked"
    assert any(f"preview_value_mismatch:{field}" in item for item in report["validation_errors"])


def test_preview_csv_row_count_mismatch_blocks(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    csv_path = root / "data/staging/bitradex_ocr/package_20260714_151816/BITRADEX_OCR_IMPORT_PREVIEW_V4.csv"
    lines = csv_path.read_text(encoding="utf-8").splitlines()
    csv_path.write_text("\n".join(lines[:-1]) + "\n", encoding="utf-8")
    report = audit(root)
    assert report["status"] == "blocked"
    assert report["preview_csv_row_count"] == 503


@pytest.mark.parametrize("kind", ["xlsx", "parquet"])
def test_master_hash_divergence_blocks(
    tmp_path: Path, project_template: Path, kind: str
) -> None:
    root = copy_project(tmp_path, project_template)
    mutate_contract(root, "expected_master_hashes", f"{kind}_sha256", value="a" * 64)
    report = audit(root)
    assert report["status"] == "blocked"
    assert any(f"master_{kind}_sha256_mismatch" in item for item in report["validation_errors"])


def test_parquet_schema_divergence_blocks(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    path = root / "data/trades/trades_master.parquet"
    frame = pd.read_parquet(path).rename(columns={"ocr_source": "unexpected"})
    frame.to_parquet(path, index=False)
    mutate_contract(root, "expected_master_hashes", "parquet_sha256", value=sha256(path))
    report = audit(root)
    assert report["status"] == "blocked"
    assert "master_parquet_schema_mismatch" in report["validation_errors"]


def test_master_row_count_divergence_blocks(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    path = root / "data/trades/trades_master.parquet"
    pd.read_parquet(path).iloc[:-1].to_parquet(path, index=False)
    mutate_contract(root, "expected_master_hashes", "parquet_sha256", value=sha256(path))
    report = audit(root)
    assert report["status"] == "blocked"
    assert report["master_parquet_row_count"] == 3057


def test_xlsx_row_count_divergence_blocks(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    path = root / "data/trades/trades_master.xlsx"
    workbook = load_workbook(path)
    worksheet = workbook["trades_master_candidate"]
    worksheet.delete_rows(worksheet.max_row)
    workbook.save(path)
    workbook.close()
    mutate_contract(root, "expected_master_hashes", "xlsx_sha256", value=sha256(path))
    report = audit(root)
    assert report["status"] == "blocked"
    assert report["master_xlsx_row_count"] == 3057


def test_no_write_does_not_create_reports(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    report = audit(root)
    assert report["write_performed"] is False
    assert not (root / "data/reports").exists()


def test_write_report_creates_only_json_and_markdown(
    tmp_path: Path, project_template: Path
) -> None:
    root = copy_project(tmp_path, project_template)
    before = {path.relative_to(root) for path in root.rglob("*") if path.is_file()}
    report = audit(root, write_report=True)
    after = {path.relative_to(root) for path in root.rglob("*") if path.is_file()}
    assert report["write_performed"] is True
    assert after - before == {
        Path("data/reports/bitradex_ocr_legacy_compatibility_v1.json"),
        Path("data/reports/bitradex_ocr_legacy_compatibility_v1.md"),
    }


def test_symlink_input_is_rejected(
    tmp_path: Path, project_template: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    root = copy_project(tmp_path, project_template)
    target = root / "data/trades/trades_master.parquet"
    original = Path.is_symlink

    def fake_is_symlink(path: Path) -> bool:
        return path == target or original(path)

    monkeypatch.setattr(Path, "is_symlink", fake_is_symlink)
    report = audit(root)
    assert report["status"] == "blocked"
    assert "master_parquet:input_symlink_rejected" in report["validation_errors"]


def test_cli_json_executes(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    completed = subprocess.run(
        [sys.executable, str(SCRIPT), "--project-root", str(root), "--no-write", "--json"],
        check=True,
        capture_output=True,
        text=True,
    )
    report = json.loads(completed.stdout)
    assert report["status"] == "ok"
    assert report["write_performed"] is False


def test_report_contains_all_safety_flags(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    report = audit(root)
    assert report["schema_version"] == AUDIT_SCHEMA_VERSION
    assert report["safety_flags"] == SAFETY_FLAGS
    for field, expected in SAFETY_FLAGS.items():
        assert report[field] is expected


def test_source_contains_no_prohibited_import_or_write_calls() -> None:
    source = (
        ROOT
        / "smartcrypto/data/bitradex_ocr_legacy_compatibility/compatibility_audit.py"
    ).read_text(encoding="utf-8")
    prohibited = (
        "build_bitradex_ocr_readonly_adapter_report(",
        "validate_staging_records(",
        ".to_excel(",
        ".to_parquet(",
        "sqlite3.connect(",
        "shutil.copyfile(",
        "subprocess",
    )
    assert all(token not in source for token in prohibited)


def test_financial_adapter_v2_remains_unchanged() -> None:
    path = ROOT / "smartcrypto/data/trader_master_fingerprint_v2/bitradex_ocr_adapter.py"
    assert sha256(path) == ADAPTER_V2_SHA256


def test_case_and_format_only_key_normalization_is_supported(
    tmp_path: Path, project_template: Path
) -> None:
    root = copy_project(tmp_path, project_template)
    summary = root / "data/staging/bitradex_ocr/package_20260714_151816/BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json"
    payload = valid_preview()
    payload["Incoming Rows"] = payload.pop("incoming_rows")
    summary.write_text(json.dumps(payload), encoding="utf-8")
    report = audit(root)
    assert report["status"] == "ok"
    assert report["preview_field_evidence"]["incoming_rows"]["source_path"] == "Incoming Rows"


def test_missing_preview_value_is_not_inferred(tmp_path: Path, project_template: Path) -> None:
    root = copy_project(tmp_path, project_template)
    summary = root / "data/staging/bitradex_ocr/package_20260714_151816/BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json"
    payload = valid_preview()
    del payload["incoming_rows"]
    payload["unrelated_total"] = 504
    summary.write_text(json.dumps(payload), encoding="utf-8")
    report = audit(root)
    assert report["status"] == "blocked"
    assert "preview_required_field_missing:incoming_rows" in report["validation_errors"]
