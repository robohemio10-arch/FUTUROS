from __future__ import annotations

import csv
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from smartcrypto.data.bitradex_ocr_legacy_authorized_append import (
    executor as v1_executor,
)
from smartcrypto.data.bitradex_ocr_legacy_authorized_append.contract import (
    parse_transition_contract as parse_v1_transition_contract,
)
from smartcrypto.data.bitradex_ocr_legacy_canonical_xlsx_transition_v2.contract import (
    AUTHORIZATION_PHRASE,
    IMPORTED_AT_UTC,
    file_sha256,
)
from smartcrypto.data.bitradex_ocr_legacy_canonical_xlsx_transition_v2.executor import (
    apply_canonical_xlsx_transition,
    verify_canonical_xlsx_transition,
)
from smartcrypto.data.bitradex_ocr_legacy_canonical_xlsx_transition_v2.planner import (
    build_canonical_xlsx_transition_plan,
    materialize_candidate_imported_at,
    semantic_rows_sha256,
)
from smartcrypto.data.trader_master_fingerprint_v2 import (
    legacy_master_governance as governance,
)


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = (
    ROOT
    / "scripts/build_bitradex_ocr_legacy_canonical_xlsx_transition_v2.py"
)
V1_CONFIG = ROOT / "config/bitradex_ocr_legacy_append_transition_v1.json"
V2_CONFIG = (
    ROOT
    / "config/bitradex_ocr_legacy_canonical_xlsx_transition_v2.json"
)
SOURCE_CONTRACT = ROOT / "config/bitradex_ocr_legacy_contract_v1.json"
POLICY = ROOT / "config/trader_master_legacy_research_only_policy_v1.json"
PACKAGE_FILES = (
    "__init__.py",
    "contract.py",
    "planner.py",
    "transaction.py",
    "executor.py",
)


def make_row(index: int, *, candidate: bool) -> dict[str, str]:
    prefix = "candidate" if candidate else "master"
    return {
        "moeda": "BTCUSDT" if index % 2 else "ETHUSDT",
        "fechar_side": "long" if index % 3 else "short",
        "leverage": "2",
        "order_id": f"{prefix}-order-{index}",
        "pnl_fechado": str(index / 100),
        "taxa_lucros_perdas_fechados_pct": str(index / 10000),
        "preco_abertura": str(60000 + index),
        "preco_fechamento": str(60001 + index),
        "volume_posicao": "0.1",
        "volume_fechado": "0.1",
        "horario_abertura": (
            f"2026-06-{(index % 28) + 1:02d}T00:00:00Z"
        ),
        "horario_fechamento": (
            f"2026-06-{(index % 28) + 1:02d}T00:05:00Z"
        ),
        "taxa_1": "-0.01",
        "preco_transacao": str(60001 + index),
        "volume_transacao": "0.1",
        "direcao_liquidez": "TAKER",
        "taxa_2": "-0.01",
        "horario_transacao": (
            f"2026-06-{(index % 28) + 1:02d}T00:05:00Z"
        ),
        "source_file": (
            "fixture_candidate_package"
            if candidate
            else f"master-{index}.jpg"
        ),
        "imported_at": (
            "" if candidate else "2026-07-14T12:00:00+00:00"
        ),
        "_dedup_key": f"{prefix}-dedup-{index}",
        "_relaxed_dedup_key": f"{prefix}-relaxed-{index}",
        "exchange_source": "BITRADEX",
        "market_data_source": "BINANCE_FUTURES",
        "ocr_source": "fixture_ocr",
    }


def write_legacy_xlsx(
    path: Path,
    header: list[str],
    *,
    include_summary: bool = True,
) -> None:
    workbook = Workbook(write_only=True)
    data_sheet = workbook.create_sheet("trades_master_candidate")
    data_sheet.append(header)
    for index in range(3058):
        data_sheet.append(
            [f"image-{index}.png", *([None] * (len(header) - 1))]
        )
    if include_summary:
        summary = workbook.create_sheet("BUILD_SUMMARY")
        summary.append(["metric", "value"])
        summary.append(["rows", 3058])
    workbook.save(path)


def copy_sources(root: Path) -> dict[str, str]:
    destination = (
        root
        / "smartcrypto/data/"
        "bitradex_ocr_legacy_canonical_xlsx_transition_v2"
    )
    destination.mkdir(parents=True)
    hashes: dict[str, str] = {}
    for name in PACKAGE_FILES:
        relative = (
            "smartcrypto/data/"
            "bitradex_ocr_legacy_canonical_xlsx_transition_v2/"
            f"{name}"
        )
        shutil.copy2(ROOT / relative, root / relative)
        hashes[relative] = file_sha256(root / relative)
    cli = (
        "scripts/"
        "build_bitradex_ocr_legacy_canonical_xlsx_transition_v2.py"
    )
    (root / "scripts").mkdir()
    shutil.copy2(ROOT / cli, root / cli)
    hashes[cli] = file_sha256(root / cli)
    return hashes


@pytest.fixture(scope="module")
def template(tmp_path_factory: pytest.TempPathFactory) -> Path:
    root = tmp_path_factory.mktemp("canonical-xlsx-v2-template")
    trades = root / "data/trades"
    package = root / "data/staging/bitradex_ocr/package_20260714_151816"
    config = root / "config"
    trades.mkdir(parents=True)
    package.mkdir(parents=True)
    config.mkdir(parents=True)

    source_payload = json.loads(SOURCE_CONTRACT.read_text(encoding="utf-8"))
    columns = source_payload["historical_master_schema"]["columns"]
    master_rows = [
        make_row(index, candidate=False) for index in range(3058)
    ]
    candidate_rows = [
        make_row(index, candidate=True) for index in range(504)
    ]
    pd.DataFrame(master_rows, columns=columns).to_parquet(
        trades / "trades_master.parquet", index=False
    )
    transition = json.loads(V2_CONFIG.read_text(encoding="utf-8"))
    raw_header = transition["pre_state"]["legacy_data_header"]
    write_legacy_xlsx(trades / "trades_master.xlsx", raw_header)

    source_payload["expected_master_hashes"] = {
        "xlsx_sha256": file_sha256(trades / "trades_master.xlsx"),
        "parquet_sha256": file_sha256(trades / "trades_master.parquet"),
    }
    (config / SOURCE_CONTRACT.name).write_text(
        json.dumps(source_payload), encoding="utf-8"
    )
    (package / "BITRADEX_OCR_IMPORT_PREVIEW_V4_SUMMARY.json").write_text(
        json.dumps(
            {
                "status": "ok",
                "incoming_rows": 504,
                "import_executed": False,
                "master_preserved": True,
            }
        ),
        encoding="utf-8",
    )
    imported_source = package / "ORDERID_SYNTHETIC_V5_SUMMARY.json"
    imported_source.write_text(
        json.dumps({"finalized_at_utc": IMPORTED_AT_UTC}),
        encoding="utf-8",
    )
    preview = package / "BITRADEX_OCR_IMPORT_PREVIEW_V4.csv"
    with preview.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[*columns, "preview_v4_classification"],
        )
        writer.writeheader()
        for row in candidate_rows:
            writer.writerow(
                {
                    **row,
                    "preview_v4_classification": (
                        "NOVEL_TO_RECONCILED_MASTER"
                    ),
                }
            )

    materialized, _ = materialize_candidate_imported_at(
        candidate_rows, columns, IMPORTED_AT_UTC
    )
    transition["pre_state"]["master_xlsx_sha256"] = file_sha256(
        trades / "trades_master.xlsx"
    )
    transition["pre_state"]["master_parquet_sha256"] = file_sha256(
        trades / "trades_master.parquet"
    )
    transition["target_state"]["expected_prefix_semantic_sha256"] = (
        semantic_rows_sha256(master_rows, columns)
    )
    transition["target_state"]["expected_tail_semantic_sha256"] = (
        semantic_rows_sha256(materialized, columns)
    )
    transition["target_state"]["expected_target_semantic_sha256"] = (
        semantic_rows_sha256([*master_rows, *materialized], columns)
    )
    transition["imported_at_policy"]["source_file_sha256"] = file_sha256(
        imported_source
    )
    transition["authorized_source_sha256"] = copy_sources(root)
    (config / V2_CONFIG.name).write_text(
        json.dumps(transition), encoding="utf-8"
    )

    policy = json.loads(POLICY.read_text(encoding="utf-8"))
    policy.update(
        expected_sha256=file_sha256(trades / "trades_master.parquet"),
        expected_size_bytes=(trades / "trades_master.parquet").stat().st_size,
        expected_row_count=3058,
        expected_schema_columns=columns,
    )
    (config / POLICY.name).write_text(
        json.dumps(policy), encoding="utf-8"
    )
    return root


def project(tmp_path: Path, template: Path) -> Path:
    root = tmp_path / "project"
    shutil.copytree(template, root)
    return root


def transition_path(root: Path) -> Path:
    return root / "config" / V2_CONFIG.name


def update_transition_hashes(root: Path) -> None:
    path = transition_path(root)
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["pre_state"]["master_xlsx_sha256"] = file_sha256(
        root / "data/trades/trades_master.xlsx"
    )
    payload["pre_state"]["master_parquet_sha256"] = file_sha256(
        root / "data/trades/trades_master.parquet"
    )
    path.write_text(json.dumps(payload), encoding="utf-8")


def hashes(root: Path) -> tuple[str, str]:
    return (
        file_sha256(root / "data/trades/trades_master.xlsx"),
        file_sha256(root / "data/trades/trades_master.parquet"),
    )


def plan(root: Path) -> dict[str, Any]:
    return build_canonical_xlsx_transition_plan(project_root=root)


def test_v1_is_definitively_blocked_without_lock_or_backup(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    root = tmp_path / "project"
    backup_directory = (
        root
        / "data"
        / "backups"
        / "bitradex_ocr_legacy_append"
    )
    backup_directory.mkdir(parents=True)

    versioned_payload = json.loads(
        V1_CONFIG.read_text(encoding="utf-8")
    )
    contract = parse_v1_transition_contract(versioned_payload)
    assert contract.transition_state == "failed_pre_replace_superseded"

    monkeypatch.setattr(
        v1_executor,
        "load_transition_contract",
        lambda *_args, **_kwargs: contract,
    )

    sentinel = backup_directory / "preexisting-backup-evidence.txt"
    sentinel.write_text(
        "preserve-existing-backup-evidence\n",
        encoding="utf-8",
    )

    before_backups = {
        path.relative_to(backup_directory).as_posix(): file_sha256(path)
        for path in backup_directory.rglob("*")
        if path.is_file()
    }

    report = v1_executor.apply_authorized_append(
        project_root=root,
        transition_contract_path=V1_CONFIG,
        expected_plan_sha256=None,
        authorization_phrase=None,
    )

    after_backups = {
        path.relative_to(backup_directory).as_posix(): file_sha256(path)
        for path in backup_directory.rglob("*")
        if path.is_file()
    }

    assert report["status"] == "blocked"
    assert (
        report["reason"]
        == "transition_v1_superseded_after_xlsx_layout_mismatch"
    )
    assert report["apply_executed"] is False
    assert report["backup_created"] is False
    assert report["write_performed"] is False
    assert report["writes_trader_master"] is False

    assert before_backups == after_backups
    assert before_backups == {
        "preexisting-backup-evidence.txt": file_sha256(sentinel)
    }

    assert not (
        root
        / "data"
        / "locks"
        / "bitradex_ocr_legacy_append_v1.lock"
    ).exists()


def test_real_legacy_layout_is_detected(template: Path) -> None:
    report = plan(template)
    assert report["status"] == "ok"
    assert report["xlsx_pre_layout_verified"] is True
    assert (
        report["pre_xlsx_classification"]
        == "legacy_ocr_evidence_workbook"
    )
    assert report["legacy_xlsx_layout"]["header_column_count"] == 71


@pytest.mark.parametrize("column_index", [0, 20, 70])
def test_any_raw_header_divergence_blocks(
    tmp_path: Path,
    template: Path,
    column_index: int,
) -> None:
    root = project(tmp_path, template)
    payload = json.loads(transition_path(root).read_text(encoding="utf-8"))
    payload["pre_state"]["legacy_data_header"][column_index] += "_drift"
    transition_path(root).write_text(json.dumps(payload), encoding="utf-8")
    report = plan(root)
    assert report["status"] == "blocked"
    assert "legacy_xlsx_header_mismatch" in report["validation_errors"]


def test_missing_build_summary_blocks(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    xlsx = root / "data/trades/trades_master.xlsx"
    workbook = load_workbook(xlsx)
    del workbook["BUILD_SUMMARY"]
    workbook.save(xlsx)
    workbook.close()
    update_transition_hashes(root)
    report = plan(root)
    assert "legacy_xlsx_build_summary_missing" in report["validation_errors"]


def test_parquet_schema_divergence_blocks(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    parquet = root / "data/trades/trades_master.parquet"
    frame = pd.read_parquet(parquet).drop(columns=["ocr_source"])
    frame.to_parquet(parquet, index=False)
    update_transition_hashes(root)
    report = plan(root)
    assert report["status"] == "blocked"
    assert "master_parquet_schema_mismatch" in report["validation_errors"]


def test_plan_builds_canonical_targets_without_project_writes(
    template: Path,
) -> None:
    before = hashes(template)
    report = plan(template)
    assert report["status"] == "ok"
    assert report["canonical_xlsx_target_ready"] is True
    assert report["target_rows"] == 3562
    assert report["xlsx_row_count"] == 3562
    assert report["parquet_row_count"] == 3562
    assert report["target_schema_column_count"] == 25
    assert report["cross_format_semantic_equality"] is True
    assert report["planner_temporary_artifacts_removed"] is True
    assert hashes(template) == before
    assert report["master_write_performed"] is False


def test_successful_fixture_apply_materializes_exact_prefix_tail_and_formats(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    current_plan = plan(root)
    result = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
    )
    assert result["status"] == "ok"
    verification = verify_canonical_xlsx_transition(project_root=root)
    assert verification["status"] == "ok"
    assert verification["master_xlsx_row_count"] == 3562
    assert verification["master_parquet_row_count"] == 3562
    assert verification["schema_column_count"] == 25
    assert verification["cross_format_semantic_equality"] is True


def test_backup_is_byte_exact_and_pre_replace_failure_persists_report(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    before = hashes(root)
    current_plan = plan(root)

    def fail(stage: str) -> None:
        if stage == "targets_verified":
            raise RuntimeError("synthetic_pre_replace_failure")

    result = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
        fault_hook=fail,
    )
    assert result["status"] == "blocked"
    assert result["failed_stage"] == "targets_verified"
    assert result["master_replace_started"] is False
    assert result["rollback_attempted"] is False
    assert result["masters_preserved"] is True
    assert hashes(root) == before
    assert result["backup_hashes"] == {
        "xlsx": before[0],
        "parquet": before[1],
    }
    assert result["report_write_performed"] is True
    assert (
        root
        / "data/reports/"
        "bitradex_ocr_legacy_canonical_xlsx_transition_v2.json"
    ).is_file()
    assert not (
        root
        / "data/locks/"
        "bitradex_ocr_legacy_canonical_xlsx_transition_v2.lock"
    ).exists()


def test_failure_between_replaces_rolls_back_both_masters(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    before = hashes(root)
    current_plan = plan(root)

    def fail(stage: str) -> None:
        if stage == "parquet_replaced":
            raise RuntimeError("synthetic_between_replace_failure")

    result = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
        fault_hook=fail,
    )
    assert result["status"] == "blocked"
    assert result["failed_stage"] == "parquet_replace"
    assert result["master_replace_started"] is True
    assert result["rollback_attempted"] is True
    assert result["rollback_succeeded"] is True
    assert result["masters_preserved"] is True
    assert hashes(root) == before
    assert result["after_failure_hashes"] == {
        "xlsx": before[0],
        "parquet": before[1],
    }
    assert result["report_write_performed"] is True
    assert not (
        root
        / "data/locks/"
        "bitradex_ocr_legacy_canonical_xlsx_transition_v2.lock"
    ).exists()


def test_reapply_is_blocked_after_success(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    current_plan = plan(root)
    first = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
    )
    second = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
    )
    assert first["status"] == "ok"
    assert second["status"] == "blocked"
    assert second["reason"] == "recomputed_plan_blocked"


def test_pre_and_post_transition_policy_are_valid(
    tmp_path: Path, template: Path
) -> None:
    root = project(tmp_path, template)
    current_plan = plan(root)
    assert current_plan["pre_transition_policy_valid"] is True
    assert current_plan["post_transition_target_policy_valid"] is True
    result = apply_canonical_xlsx_transition(
        project_root=root,
        expected_plan_sha256=current_plan["plan_sha256"],
        authorization_phrase=AUTHORIZATION_PHRASE,
    )
    assert result["status"] == "ok"
    evidence = governance._authorized_transition_target_artifact_evidence(
        root,
        trader_master_path="data/trades/trades_master.parquet",
        transition_contract_path=transition_path(root),
    )
    assert evidence["post_transition_artifact_matches"] is True


def test_governance_recognizes_hash_pinned_v2_sources(
    template: Path,
) -> None:
    paths = tuple(
        sorted(
            path.relative_to(template).as_posix()
            for path in template.rglob("*")
            if path.is_file()
        )
    )
    evidence = governance._guarded_transition_evidence(
        template,
        transition_contract_path=transition_path(template),
        tracked_paths=paths,
    )
    assert evidence["guarded_transition_source_hashes_valid"] is True
    assert evidence["pre_transition_policy_valid"] is True
    assert evidence["post_transition_target_policy_valid"] is True


def test_zero_operational_authority_and_no_external_capabilities(
    template: Path,
) -> None:
    report = plan(template)
    for field in (
        "operational_authority",
        "apply_executed",
        "import_executed",
        "master_write_performed",
        "writes_trader_master",
        "writes_xlsx",
        "writes_parquet",
        "writes_sqlite",
        "writes_runtime",
        "changes_risk",
        "sends_orders",
        "exchange_private_access",
        "updates_qlib",
        "updates_ai_shadow",
    ):
        assert report[field] is False
    source = "\n".join(
        (ROOT / relative).read_text(encoding="utf-8").casefold()
        for relative in (
            "smartcrypto/data/"
            "bitradex_ocr_legacy_canonical_xlsx_transition_v2/executor.py",
            "smartcrypto/data/"
            "bitradex_ocr_legacy_canonical_xlsx_transition_v2/planner.py",
        )
    )
    for token in (
        "import ccxt",
        "from ccxt",
        "import requests",
        "import socket",
        "riskmanager",
        "send_order",
    ):
        assert token not in source


def test_cli_plan_executes_without_apply(template: Path) -> None:
    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "plan",
            "--project-root",
            str(template),
            "--json",
        ],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    )
    payload = json.loads(completed.stdout)
    assert payload["status"] == "ok"
    assert payload["apply_executed"] is False
    assert payload["master_write_performed"] is False


def test_transition_contract_pins_expected_pre_state() -> None:
    payload = json.loads(V2_CONFIG.read_text(encoding="utf-8"))
    pre_state = payload["pre_state"]
    target_state = payload["target_state"]

    assert payload["transition_state"] == "planned_not_executed"

    assert pre_state["master_xlsx_sha256"] == (
        "83e2d17db317cc84b2bd39e00a961bd8d568c4375c5a4a113f6a26df58972e90"
    )
    assert pre_state["master_parquet_sha256"] == (
        "24e049b3ca7a72dbde071a056548035fed87651d48959cd0cf4c6c8b0dac7295"
    )

    assert pre_state["master_row_count"] == 3058
    assert pre_state["canonical_schema_column_count"] == 25
    assert (
        pre_state["xlsx_classification"]
        == "legacy_ocr_evidence_workbook"
    )
    assert (
        pre_state["legacy_data_sheet"]
        == "trades_master_candidate"
    )
    assert pre_state["legacy_summary_sheet"] == "BUILD_SUMMARY"

    assert target_state["expected_row_count"] == 3562
    assert target_state["canonical_schema_column_count"] == 25


def test_new_transition_uses_distinct_authority() -> None:
    v1 = json.loads(V1_CONFIG.read_text(encoding="utf-8"))
    v2 = json.loads(V2_CONFIG.read_text(encoding="utf-8"))
    assert v1["transition_state"] == "failed_pre_replace_superseded"
    assert v1["transition_id"] != v2["transition_id"]
    assert v1["schema_version"] != v2["schema_version"]
    assert (
        v1["execution_policy"]["authorization_phrase"]
        != v2["execution_policy"]["authorization_phrase"]
    )
    assert hashlib.sha256(V1_CONFIG.read_bytes()).hexdigest() != hashlib.sha256(
        V2_CONFIG.read_bytes()
    ).hexdigest()
